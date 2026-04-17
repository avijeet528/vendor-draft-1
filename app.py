# ============================================================
# app.py — IT Procurement Intelligence Dashboard (Cleaned)
# Part 1 of continuation
# ============================================================

import io
import os
import re
import zipfile
from collections import defaultdict

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

try:
    import requests
    REQUESTS_OK = True
except ImportError:
    REQUESTS_OK = False

try:
    import pdfplumber
    PDF_OK = True
except ImportError:
    PDF_OK = False


# ============================================================
# APP CONFIG
# ============================================================

st.set_page_config(
    page_title="IT Procurement Intelligence",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ============================================================
# THEME
# ============================================================

C_ORANGE = "#D04A02"
C_ORANGE_DARK = "#A33A00"
C_ORANGE_MID = "#E8703A"
C_ORANGE_LITE = "#FAD4C0"

C_BLACK = "#1A1A1A"
C_DARK = "#2D2D2D"
C_MID = "#4A4A4A"

C_GREY_DARK = "#7D7D7D"
C_GREY = "#B0B0B0"
C_GREY_LITE = "#E0E0E0"
C_GREY_BG = "#F3F3F3"
C_WHITE = "#FFFFFF"

CHART_SEQ = [
    C_ORANGE,
    C_DARK,
    C_ORANGE_MID,
    C_GREY_DARK,
    C_ORANGE_DARK,
    C_GREY,
    C_MID,
    C_BLACK,
]

CFONT = dict(
    family="Georgia,'Source Sans Pro',Arial",
    size=11,
    color=C_DARK,
)

CBG = C_GREY_BG
DEMO_DIR = "demo_quotes"
MASTER_CATALOG_FILE = "Master Catalog.xlsx"
DUMMY_CATALOG_FILE = "dummy_catalog.csv"


# ============================================================
# GLOBAL CSS
# ============================================================

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Source+Sans+Pro:wght@300;400;600;700&display=swap');

    html, body, [class*="css"], div, p, span, td, th, label, button, .stMarkdown {
        font-family: 'Source Sans Pro', 'Helvetica Neue', Arial, sans-serif !important;
    }

    h1, h2, h3 {
        font-family: Georgia, 'ITC Charter', serif !important;
        font-weight: 700 !important;
    }

    section[data-testid="stSidebar"] { display: none !important; }
    [data-testid="collapsedControl"] { display: none !important; }
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    header { visibility: hidden; }

    .main .block-container {
        background: #F3F3F3 !important;
        max-width: 100% !important;
        padding: 1.2rem 2.5rem !important;
    }

    button[data-baseweb="tab"] {
        font-weight: 600 !important;
        font-size: 0.90em !important;
        color: #7D7D7D !important;
    }

    button[data-baseweb="tab"][aria-selected="true"] {
        color: #D04A02 !important;
        border-bottom: 3px solid #D04A02 !important;
        background: transparent !important;
    }

    .kpi-box {
        border-radius: 4px;
        padding: 20px 12px;
        text-align: center;
        color: white;
        border-left: 5px solid rgba(255,255,255,0.2);
    }

    .kpi-value {
        font-size: 2.4em;
        font-weight: 700;
        margin: 0;
        line-height: 1.1;
        font-family: Georgia, serif !important;
    }

    .kpi-label {
        font-size: 0.72em;
        font-weight: 700;
        opacity: 0.9;
        margin-top: 6px;
        letter-spacing: 1.2px;
        text-transform: uppercase;
    }

    .sec-head {
        font-size: 0.73em;
        font-weight: 700;
        letter-spacing: 1.4px;
        text-transform: uppercase;
        color: #D04A02;
        margin: 22px 0 10px;
        border-bottom: 2px solid #D04A02;
        padding-bottom: 5px;
        display: block;
    }

    .comp-table {
        width: 100%;
        border-collapse: collapse;
        font-size: 0.81em;
        border: 1px solid #E0E0E0;
    }

    .comp-table thead tr { background: #2D2D2D; }

    .comp-table thead th {
        padding: 10px 12px;
        text-align: left;
        font-weight: 700;
        font-size: 0.78em;
        letter-spacing: 0.5px;
        text-transform: uppercase;
        color: white !important;
        border: none;
    }

    .comp-table tbody tr:nth-child(even) { background: #F8F8F8; }
    .comp-table tbody tr:hover { background: #FAD4C0; }

    .comp-table tbody td {
        padding: 9px 12px;
        border-bottom: 1px solid #EBEBEB;
        vertical-align: middle;
        word-break: break-word;
        color: #2D2D2D;
    }

    .vbadge {
        display: inline-block;
        padding: 3px 9px;
        border-radius: 2px;
        color: white;
        font-size: 0.76em;
        font-weight: 700;
        white-space: nowrap;
    }

    .scard {
        border-radius: 4px;
        padding: 16px;
        margin-bottom: 10px;
        border-left: 5px solid #D04A02;
        background: white;
    }

    .scard-orange { border-color: #D04A02; background: #FFF5F0; }
    .scard-dark { border-color: #2D2D2D; background: #F5F5F5; }
    .scard-grey { border-color: #7D7D7D; background: #FAFAFA; }

    .verdict-good {
        background: #FFF5F0;
        border: 2px solid #D04A02;
        border-radius: 4px;
        padding: 14px 18px;
        color: #D04A02;
        font-weight: 700;
    }

    .verdict-mid {
        background: #F5F5F5;
        border: 2px solid #4A4A4A;
        border-radius: 4px;
        padding: 14px 18px;
        color: #4A4A4A;
        font-weight: 700;
    }

    .verdict-bad {
        background: #F0F0F0;
        border: 2px solid #7D7D7D;
        border-radius: 4px;
        padding: 14px 18px;
        color: #2D2D2D;
        font-weight: 700;
    }

    .chat-header {
        background: #2D2D2D;
        color: white;
        padding: 12px 18px;
        border-radius: 6px 6px 0 0;
    }

    .chat-outer {
        background: #F8F8F8;
        border: 1px solid #E0E0E0;
        border-top: none;
        border-radius: 0;
        padding: 14px 14px 6px;
        min-height: 320px;
        max-height: 420px;
        overflow-y: auto;
    }

    .msg-user {
        background: #D04A02;
        color: white;
        border-radius: 14px 14px 3px 14px;
        padding: 9px 14px;
        margin: 5px 0 5px auto;
        max-width: 74%;
        font-size: 0.86em;
        display: inline-block;
        float: right;
        clear: both;
    }

    .msg-bot {
        background: white;
        color: #2D2D2D;
        border: 1px solid #E0E0E0;
        border-left: 4px solid #D04A02;
        border-radius: 14px 14px 14px 3px;
        padding: 9px 14px;
        margin: 5px 0;
        max-width: 84%;
        font-size: 0.86em;
        display: inline-block;
        float: left;
        clear: both;
    }

    .chat-wrap {
        overflow: hidden;
        margin-bottom: 3px;
    }

    .filter-bar {
        background: #2D2D2D;
        padding: 14px 20px;
        border-radius: 4px;
        margin-bottom: 18px;
        border-left: 4px solid #D04A02;
    }

    .insight-box {
        background: #FFF5F0;
        border-left: 4px solid #D04A02;
        border-radius: 0 4px 4px 0;
        padding: 11px 16px;
        margin: 10px 0;
        font-size: 0.86em;
        color: #2D2D2D;
    }

    .bucket-header {
        background: #1A1A1A;
        color: white;
        padding: 10px 20px;
        border-radius: 4px;
        margin-bottom: 6px;
        border-left: 6px solid #D04A02;
        font-size: 0.85em;
        font-weight: 700;
        letter-spacing: 1px;
        text-transform: uppercase;
    }

    div[data-testid="stHorizontalBlock"] button[kind="secondary"] {
        background: white !important;
        border: 1.5px solid #D04A02 !important;
        color: #D04A02 !important;
        border-radius: 20px !important;
        font-size: 0.78em !important;
        font-weight: 600 !important;
        padding: 4px 8px !important;
    }

    div[data-testid="stHorizontalBlock"] button[kind="secondary"]:hover {
        background: #FFF5F0 !important;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ============================================================
# REGEX / PRICE EXTRACTION
# ============================================================

PRICE_RE = re.compile(
    r"(?:USD|EUR|GBP|SGD|MYR|AUD|CAD)\s?\d{1,3}(?:[,]\d{3})*(?:\.\d{1,2})?"
    r"|(?:[\$\€\£]\s?)\d{1,3}(?:[,\s]\d{3})*(?:\.\d{1,2})?"
    r"|\d{1,3}(?:[,]\d{3})+(?:\.\d{1,2})?",
    re.IGNORECASE,
)

TOTAL_KW = [
    "grand total",
    "total amount",
    "total price",
    "amount due",
    "net total",
    "total cost",
    "total value",
    "subtotal",
    "total",
]


# ============================================================
# GENERIC HELPERS
# ============================================================

def parse_num(value) -> float:
    try:
        return float(re.sub(r"[^\d.]", "", str(value)) or "0")
    except Exception:
        return 0.0


def fmt_currency(value) -> str:
    try:
        parsed = float(re.sub(r"[^\d.]", "", str(value)) or "0")
        return "—" if parsed <= 0 else "${:,.2f}".format(parsed)
    except Exception:
        return str(value)


def best_price_from_text(text: str) -> str:
    text_lower = text.lower()

    for keyword in TOTAL_KW:
        idx = text_lower.find(keyword)
        if idx == -1:
            continue

        snippet = text[max(0, idx - 20): idx + 300]
        hits = PRICE_RE.findall(snippet)
        valid = [hit.strip() for hit in hits if parse_num(hit) >= 50]
        if valid:
            return max(valid, key=parse_num)

    all_hits = PRICE_RE.findall(text)
    valid = [hit.strip() for hit in all_hits if parse_num(hit) >= 100]
    return max(valid, key=parse_num) if valid else ""


def text_from_bytes(content: bytes, ext: str) -> str:
    text = ""
    ext = ext.lower().strip(".")

    try:
        if ext == "pdf":
            if not PDF_OK:
                return ""
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"

        elif ext in ("xlsx", "xls"):
            workbook = openpyxl.load_workbook(
                io.BytesIO(content),
                data_only=True,
                read_only=True,
            )
            rows_text = []
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(values_only=True):
                    row_text = "  ".join(str(cell) for cell in row if cell is not None)
                    if row_text.strip():
                        rows_text.append(row_text)
            text = "\n".join(rows_text)
            workbook.close()

        elif ext == "docx":
            with zipfile.ZipFile(io.BytesIO(content)) as zipped:
                if "word/document.xml" in zipped.namelist():
                    xml = zipped.read("word/document.xml").decode("utf-8", errors="ignore")
                    text = re.sub(r"<[^>]+>", " ", xml)
                    text = re.sub(r"\s{2,}", "\n", text)

    except Exception:
        pass

    return text


def extract_price_from_bytes(content: bytes, ext: str) -> dict:
    text = text_from_bytes(content, ext)
    price = best_price_from_text(text)

    if not price or parse_num(price) <= 0:
        all_numbers = [
            hit.strip()
            for hit in PRICE_RE.findall(text)
            if parse_num(hit) >= 1000
        ]
        if all_numbers:
            price = max(all_numbers, key=parse_num)

    return {
        "price": price,
        "price_num": parse_num(price) if price else 0.0,
        "text": text[:5000],
    }


def extract_price_from_file(filepath: str) -> dict:
    try:
        with open(filepath, "rb") as file:
            content = file.read()
        ext = filepath.rsplit(".", 1)[-1]
        return extract_price_from_bytes(content, ext)
    except Exception:
        return {"price": "", "price_num": 0.0, "text": ""}


# ============================================================
# SCORING HELPERS
# ============================================================

def price_score(new_price: float, historical_prices: list[float]):
    valid = [price for price in historical_prices if price > 0]

    if not valid or new_price <= 0:
        return None, "No comparison data", 0, 0, 0

    minimum = min(valid)
    maximum = max(valid)
    average = sum(valid) / len(valid)

    if maximum == minimum:
        return 50, "Same as historical average", average, minimum, maximum

    score = round((1 - (new_price - minimum) / (maximum - minimum)) * 100, 1)
    score = max(0, min(100, score))

    pct = round((new_price - average) / average * 100, 1)
    label = (
        f"{abs(pct)}% BELOW average — COMPETITIVE"
        if new_price < average
        else f"{abs(pct)}% ABOVE average — REVIEW NEEDED"
    )

    return score, label, average, minimum, maximum


def score_color(score):
    if score is None:
        return C_GREY_DARK
    if score >= 70:
        return C_ORANGE
    if score >= 40:
        return C_GREY_DARK
    return C_MID


def get_verdict(score):
    if score is None:
        return "⚪ No Data", "No comparison data.", C_GREY_DARK
    if score >= 70:
        return "✅ COMPETITIVE", "Priced competitively.", C_ORANGE
    if score >= 40:
        return "🟡 AVERAGE", "Within range. Negotiate.", C_GREY_DARK
    return "🔴 HIGH", "Above average. Recommend negotiating.", C_MID


# ============================================================
# DATAFRAME HELPERS
# ============================================================

def clean_df(df: pd.DataFrame) -> pd.DataFrame:
    safe_columns = [col for col in df.columns if col != "Services List"]
    cleaned = df[safe_columns].copy()

    for col in cleaned.columns:
        try:
            cleaned[col] = cleaned[col].fillna("").apply(lambda x: str(x).strip())
        except Exception:
            cleaned[col] = ""

    mask = (
        cleaned["Category"].apply(lambda x: x in ["", "nan"])
        & cleaned["Vendor"].apply(lambda x: x in ["", "nan"])
    )
    cleaned = cleaned[~mask].copy()
    cleaned.reset_index(drop=True, inplace=True)
    return cleaned


def parse_services(value) -> list[str]:
    if not value or str(value).strip() in ["", "nan", "None"]:
        return ["(unspecified)"]

    text = str(value).replace("\\n", "\n").replace("\r\n", "\n").replace("\r", "\n")
    parts = [part.strip() for part in text.split("\n") if part.strip() and part.strip() != "nan"]

    if not parts:
        parts = [part.strip() for part in text.split(";") if part.strip()]

    if not parts and len(text) < 300:
        parts = [part.strip() for part in text.split(",") if part.strip()]

    return parts if parts else ["(unspecified)"]


def explode_services(df: pd.DataFrame):
    df_base = df.copy()
    df_base["Services List"] = df_base["Comments"].apply(parse_services)

    exploded = df_base.explode("Services List").copy()
    exploded.rename(columns={"Services List": "Service"}, inplace=True)
    exploded["Service"] = exploded["Service"].apply(lambda x: str(x).strip())

    exploded = exploded[
        ~exploded["Service"].isin(["", "(unspecified)", "nan", "None"])
    ].reset_index(drop=True)

    return df_base, exploded


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    col_map = {}

    for col in df.columns:
        normalized = str(col).lower().strip()

        if normalized == "category" and "Category" not in col_map:
            col_map["Category"] = col

        elif any(key in normalized for key in ["vendor", "supplier"]) and "Vendor" not in col_map:
            col_map["Vendor"] = col

        elif ("file name" in normalized or normalized == "filename") and "File Name" not in col_map:
            col_map["File Name"] = col

        elif any(key in normalized for key in ["file link", "file url"]) and "File Link" not in col_map:
            col_map["File Link"] = col

        elif any(key in normalized for key in ["comment", "service", "description", "scope"]) and "Comments" not in col_map:
            col_map["Comments"] = col

        elif any(key in normalized for key in ["price", "cost", "amount", "quoted"]) and "Quoted Price" not in col_map:
            col_map["Quoted Price"] = col

    df = df.rename(columns={v: k for k, v in col_map.items()})

    for required in ["Category", "Vendor", "File Name", "Comments"]:
        if required not in df.columns:
            df[required] = ""

    keep = ["Category", "Vendor", "File Name", "Comments"]
    for optional in ["File Link", "Quoted Price"]:
        if optional in df.columns:
            keep.append(optional)

    return df[[col for col in keep if col in df.columns]].copy()


def extract_hyperlink_map_from_excel(filepath: str) -> dict:
    hyperlink_map = {}

    try:
        workbook = openpyxl.load_workbook(filepath)
        worksheet = workbook.active

        file_col = None
        header_row = None

        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip().lower() == "file name":
                    file_col = cell.column
                    header_row = cell.row
                    break
            if file_col:
                break

        if file_col and header_row:
            for row in worksheet.iter_rows(min_row=header_row + 1):
                for cell in row:
                    if cell.column == file_col and cell.value and cell.hyperlink:
                        hyperlink_map[str(cell.value).strip()] = str(cell.hyperlink.target).strip()

        workbook.close()

    except Exception:
        pass

    return hyperlink_map


@st.cache_data
def load_master_catalog():
    if not os.path.exists(MASTER_CATALOG_FILE):
        return None, None

    try:
        raw = pd.read_excel(MASTER_CATALOG_FILE, engine="openpyxl", header=None)
        header_row = 0

        for idx, row in raw.iterrows():
            vals = [str(v).strip().lower() for v in row.values if pd.notna(v)]
            if any("category" in v for v in vals) and any("vendor" in v for v in vals):
                header_row = idx
                break

        df = pd.read_excel(MASTER_CATALOG_FILE, engine="openpyxl", header=header_row)
        df.columns = [str(col).strip() for col in df.columns]

        hyperlink_map = extract_hyperlink_map_from_excel(MASTER_CATALOG_FILE)

        df = normalize_columns(df)
        df = clean_df(df)
        df["Hyperlink"] = df["File Name"].map(hyperlink_map).fillna("")

        df, exploded = explode_services(df)
        return df, exploded

    except Exception as exc:
        st.warning(f"Master catalog error: {exc}")
        return None, None


@st.cache_data
def load_dummy_data():
    if not os.path.exists(DUMMY_CATALOG_FILE):
        return None, None

    try:
        df = pd.read_csv(DUMMY_CATALOG_FILE)
        df.columns = [str(col).strip() for col in df.columns]

        df = normalize_columns(df)
        df = clean_df(df)
        df["Hyperlink"] = ""

        df, exploded = explode_services(df)
        return df, exploded

    except Exception as exc:
        st.warning(f"Dummy data error: {exc}")
        return None, None


def ensure_dummy_data():
    if os.path.exists(DUMMY_CATALOG_FILE):
        return

    import random

    random.seed(42)

    vendors = [
        "NTT Data",
        "Dimension Data",
        "Telstra",
        "Optus",
        "Vocus",
        "Datacom",
    ]

    categories = {
        "Cybersecurity": [
            "Endpoint Protection",
            "SIEM Monitoring",
            "Privileged Access Mgmt",
            "Security Awareness",
            "Network Access Control",
        ],
        "Network & Telecom": [
            "Cisco Catalyst 9200-L",
            "Cisco Catalyst 9400",
            "Palo Alto NGFW",
            "SD-WAN Solution",
            "Cisco Meraki MX",
        ],
        "Hosting": [
            "VMware vSphere",
            "NetApp Storage",
            "Oracle DB License",
            "Colocation Build",
            "Backup & Recovery",
        ],
        "M365 & Power Platform": [
            "M365 E3 License",
            "M365 E5 License",
            "Power BI Premium",
            "Teams Rooms",
        ],
    }

    rows = []
    for category, services in categories.items():
        for service in services:
            vendor_count = random.randint(2, 4)
            chosen_vendors = random.sample(vendors, vendor_count)

            for vendor in chosen_vendors:
                base = random.uniform(20000, 200000)
                price = round(base * random.uniform(0.85, 1.15), 2)
                filename = "{}_{}_{}.pdf".format(
                    vendor.replace(" ", "_"),
                    service.replace(" ", "_")[:15],
                    random.randint(1000, 9999),
                )

                rows.append(
                    {
                        "Category": category,
                        "Vendor": vendor,
                        "File Name": filename,
                        "Comments": service,
                        "Quoted Price": price,
                    }
                )

    pd.DataFrame(rows).to_csv(DUMMY_CATALOG_FILE, index=False)


# ============================================================
# DOMAIN HELPERS
# ============================================================

def infer_subcategory(category, comments, filename):
    text = f"{comments} {filename}".lower()
    category_text = str(category).lower().strip()

    if "cybersecurity" in category_text:
        if any(key in text for key in ["trendmicro", "endpoint", "antivirus"]):
            return "Endpoint Protection"
        if any(key in text for key in ["cyberark", "privileged", "pam"]):
            return "Privileged Access"
        if any(key in text for key in ["knowbe4", "awareness", "phishing"]):
            return "Security Awareness"
        if any(key in text for key in ["forescout", "nac"]):
            return "Network Access Control"
        if any(key in text for key in ["siem", "splunk", "monitor"]):
            return "SIEM / Monitoring"
        return "General Security"

    if "network" in category_text or "telecom" in category_text:
        if "meraki" in text:
            return "Cisco Meraki"
        if "palo alto" in text:
            return "Palo Alto NGFW"
        if "equinix" in text:
            return "Equinix"
        if "cisco" in text:
            return "Cisco Networking"
        return "General Network"

    if "hosting" in category_text:
        if any(key in text for key in ["vmware", "vcf"]):
            return "VMware"
        if "oracle" in text:
            return "Oracle DB"
        if "netapp" in text:
            return "NetApp"
        if any(key in text for key in ["colo", "colocation"]):
            return "Colocation"
        return "General Hosting"

    if "m365" in category_text:
        return "M365 Licensing"
    if "idam" in category_text:
        return "Identity Migration"
    if "snow" in category_text:
        return "ServiceNow ITSM"
    if "summary" in category_text:
        return "Reporting"

    return str(category).strip().title()


def resolve_url(row) -> str:
    for col in ["Hyperlink", "File Link"]:
        value = str(row.get(col, "")).strip()
        if value and value not in ["", "nan"] and value.startswith("http"):
            return value

    filename = str(row.get("File Name", "")).strip()
    if filename:
        local_path = os.path.join(DEMO_DIR, filename)
        if os.path.exists(local_path):
            return local_path

    return ""


def vendor_color_maps(df_master, df_dummy):
    master_map = {}
    dummy_map = {}

    if df_master is not None and not df_master.empty:
        for idx, vendor in enumerate(sorted(df_master["Vendor"].unique())):
            master_map[vendor] = CHART_SEQ[idx % len(CHART_SEQ)]

    if df_dummy is not None and not df_dummy.empty:
        for idx, vendor in enumerate(sorted(df_dummy["Vendor"].unique())):
            dummy_map[vendor] = CHART_SEQ[idx % len(CHART_SEQ)]

    return master_map, dummy_map

# ============================================================
# PART 2 — UI HELPERS, CHATBOT, CHARTS, OVERVIEW RENDERING
# ============================================================

import re as _chat_re


# ============================================================
# UI HELPERS
# ============================================================

def vpill(vendor_name, color=None):
    return (
        "<span class='vbadge' style='background:{}'>{}</span>".format(
            color or C_DARK,
            vendor_name,
        )
    )


def sec(title, caption=""):
    st.markdown(
        "<span class='sec-head'>{}</span>".format(title),
        unsafe_allow_html=True,
    )
    if caption:
        st.markdown(
            "<div style='font-size:0.81em;color:#7D7D7D;margin:-6px 0 10px'>{}</div>".format(caption),
            unsafe_allow_html=True,
        )


def kpi_box(col, value, label, bg):
    col.markdown(
        "<div class='kpi-box' style='background:{}'>"
        "<div class='kpi-value'>{}</div>"
        "<div class='kpi-label'>{}</div>"
        "</div>".format(bg, value, label),
        unsafe_allow_html=True,
    )


def mini_kpi(col, value, label, bg, icon=""):
    col.markdown(
        "<div class='kpi-box' style='background:{};min-height:82px'>"
        "<div style='font-size:1.3em;margin-bottom:2px'>{}</div>"
        "<div class='kpi-value' style='font-size:1.3em'>{}</div>"
        "<div class='kpi-label'>{}</div>"
        "</div>".format(bg, icon, value, label),
        unsafe_allow_html=True,
    )


def insight(text):
    st.markdown(
        "<div class='insight-box'>💡 {}</div>".format(text),
        unsafe_allow_html=True,
    )


def pwc_bar(fig_obj, title="", height=320):
    fig_obj.update_layout(
        height=height,
        plot_bgcolor=CBG,
        paper_bgcolor=CBG,
        margin=dict(l=5, r=10, t=36 if title else 16, b=10),
        font=CFONT,
        yaxis=dict(showgrid=True, gridcolor=C_GREY_LITE, zeroline=False),
        bargap=0.35,
        showlegend=False,
    )
    if title:
        fig_obj.update_layout(
            title=dict(
                text=title,
                font=dict(size=12, color=C_DARK, family="Georgia,serif"),
                x=0,
                xanchor="left",
            )
        )


def render_html_table(rows_html: list[str]):
    st.markdown("".join(rows_html), unsafe_allow_html=True)


# ============================================================
# CHATBOT HELPERS
# ============================================================

def get_matching_services(all_services, message):
    matches = []
    for service in all_services:
        service_lower = service.lower()
        if service_lower in message:
            matches.append(service)
            continue

        tokens = [word for word in service_lower.split() if len(word) > 3]
        if any(word in message for word in tokens):
            matches.append(service)

    return matches[:8]


def chatbot_response(
    user_msg,
    df_master,
    df_exploded,
    uploaded_file_bytes=None,
    uploaded_file_name=None,
):
    message = user_msg.lower().strip()

    if df_master is None or df_exploded is None:
        return {"type": "text", "text": "No catalog loaded."}

    all_services = sorted(df_exploded["Service"].unique().tolist())
    all_vendors = sorted(df_master["Vendor"].unique().tolist())
    all_categories = sorted(df_master["Category"].unique().tolist())

    # --------------------------------------------------------
    # FILE UPLOAD HANDLING
    # --------------------------------------------------------
    if uploaded_file_bytes is not None and uploaded_file_name:
        ext = uploaded_file_name.rsplit(".", 1)[-1].lower()
        result = extract_price_from_bytes(uploaded_file_bytes, ext)
        price = result["price_num"]

        st.session_state["tab2_upload_price"] = price
        st.session_state["tab2_upload_fname"] = uploaded_file_name
        st.session_state["tab2_file_bytes"] = uploaded_file_bytes
        st.session_state["tab2_file_ext"] = ext
        st.session_state["chat_redirect_upload"] = True

        history_prices = []
        text_lower = result["text"].lower()

        matched_services = [
            service
            for service in all_services
            if any(word in text_lower for word in service.lower().split() if len(word) > 3)
        ][:5]

        if matched_services:
            for service in matched_services:
                service_rows = df_exploded[df_exploded["Service"] == service]
                for _, row in service_rows.iterrows():
                    quoted = parse_num(str(row.get("Quoted Price", "")).strip())
                    if quoted > 0:
                        history_prices.append(quoted)

        verdict_text = ""
        if history_prices and price > 0:
            score, label, avg_price, _, _ = price_score(price, history_prices)
            verdict_text = (
                "\n\n📊 vs {} similar: Avg {} | **{}**".format(
                    len(history_prices),
                    fmt_currency(avg_price),
                    label,
                )
            )

        return {
            "type": "redirect_upload",
            "text": (
                "📄 **{}** — price **{}**{}\n\n"
                "🔄 Redirecting to Upload & Score…".format(
                    uploaded_file_name,
                    fmt_currency(price) if price > 0 else "not found",
                    verdict_text,
                )
            ),
            "price": price,
        }

    # --------------------------------------------------------
    # GREETING
    # --------------------------------------------------------
    if any(word in message for word in ["hello", "hi", "hey"]):
        return {
            "type": "text",
            "text": (
                "👋 Hello! I know the full catalog.\n\n"
                "**{} quotes · {} vendors · {} services**\n\n"
                "Ask:\n"
                "• *Who quoted Cisco Catalyst?*\n"
                "• *Compare Palo Alto prices*\n"
                "• *Cheapest Cybersecurity vendor?*\n"
                "• Upload a quote file for instant scoring".format(
                    len(df_master),
                    df_master["Vendor"].nunique(),
                    df_exploded["Service"].nunique(),
                )
            ),
        }

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------
    if any(word in message for word in ["summary", "overview", "how many", "total", "catalog"]):
        lines = []
        for category in all_categories:
            category_df = df_master[df_master["Category"] == category]
            service_count = df_exploded[df_exploded["Category"] == category]["Service"].nunique()

            prices = []
            if "Quoted Price" in category_df.columns:
                for price in category_df["Quoted Price"]:
                    val = parse_num(str(price))
                    if val > 0:
                        prices.append(val)

            price_text = "avg {}".format(fmt_currency(sum(prices) / len(prices))) if prices else "no prices"

            lines.append(
                "• **{}**: {} quotes · {} vendors · {} svcs · {}".format(
                    category,
                    len(category_df),
                    category_df["Vendor"].nunique(),
                    service_count,
                    price_text,
                )
            )

        return {
            "type": "text",
            "text": (
                "📊 **Catalog:**\n"
                "{} quotes · {} vendors · {} services · {} cats\n\n{}".format(
                    len(df_master),
                    df_master["Vendor"].nunique(),
                    df_exploded["Service"].nunique(),
                    df_master["Category"].nunique(),
                    "\n".join(lines),
                )
            ),
        }

    # --------------------------------------------------------
    # PRICE ANALYSIS
    # --------------------------------------------------------
    if any(
        word in message
        for word in [
            "cheapest",
            "expensive",
            "best price",
            "competitive",
            "price analysis",
            "which vendor",
        ]
    ):
        matched_category = next(
            (
                category
                for category in all_categories
                if category.lower() in message
                or any(word in message for word in category.lower().split() if len(word) > 3)
            ),
            None,
        )

        scope_df = df_master[df_master["Category"] == matched_category] if matched_category else df_master

        vendor_avgs = {}
        for vendor in scope_df["Vendor"].unique():
            vendor_df = scope_df[scope_df["Vendor"] == vendor]
            prices = []

            if "Quoted Price" in vendor_df.columns:
                for quoted in vendor_df["Quoted Price"]:
                    value = parse_num(str(quoted))
                    if value > 0:
                        prices.append(value)

            if prices:
                vendor_avgs[vendor] = sum(prices) / len(prices)

        if vendor_avgs:
            best_vendor = min(vendor_avgs, key=vendor_avgs.get)
            worst_vendor = max(vendor_avgs, key=vendor_avgs.get)
            overall_avg = sum(vendor_avgs.values()) / len(vendor_avgs)

            lines = []
            for vendor, avg_price in sorted(vendor_avgs.items(), key=lambda x: x[1]):
                pct = round((avg_price - overall_avg) / overall_avg * 100, 1) if overall_avg > 0 else 0
                tag = (
                    "🟠 Cheapest" if vendor == best_vendor
                    else "⚫ Most exp." if vendor == worst_vendor
                    else "⚪ Mid"
                )
                lines.append(
                    "**{}**: {} ({:+.1f}%) {}".format(
                        vendor,
                        fmt_currency(avg_price),
                        pct,
                        tag,
                    )
                )

            return {
                "type": "price_analysis",
                "text": (
                    "💰 **Price Analysis{}**\n\nAvg: {}\n\n{}\n\n"
                    "✅ **{}** cheapest at **{}**".format(
                        " — " + matched_category if matched_category else "",
                        fmt_currency(overall_avg),
                        "\n".join(lines),
                        best_vendor,
                        fmt_currency(vendor_avgs[best_vendor]),
                    )
                ),
                "vendor_avgs": vendor_avgs,
                "avg": overall_avg,
                "best_vendor": best_vendor,
            }

        return {"type": "text", "text": "No price data available."}

    matched_vendor = next((vendor for vendor in all_vendors if vendor.lower() in message), None)
    matched_services = get_matching_services(all_services, message)
    matched_category = next(
        (
            category
            for category in all_categories
            if category.lower() in message
            or any(word in message for word in category.lower().split() if len(word) > 4)
        ),
        None,
    )

    # --------------------------------------------------------
    # VENDOR + SERVICE
    # --------------------------------------------------------
    if matched_vendor and matched_services:
        service = matched_services[0]
        filtered = df_exploded[
            (df_exploded["Vendor"] == matched_vendor)
            & (df_exploded["Service"] == service)
        ]

        if not filtered.empty:
            files = filtered["File Name"].unique().tolist()
            prices = []

            for filename in files:
                rows = df_master[df_master["File Name"] == filename]
                if len(rows) > 0 and "Quoted Price" in rows.columns:
                    quoted = parse_num(str(rows["Quoted Price"].values[0]))
                    if quoted > 0:
                        prices.append(quoted)

            price_text = "Price: **{}**".format(fmt_currency(prices[0])) if prices else "No price on record."

            all_prices = [
                parse_num(str(row.get("Quoted Price", "")).strip())
                for _, row in df_exploded[df_exploded["Service"] == service].iterrows()
                if parse_num(str(row.get("Quoted Price", "")).strip()) > 0
            ]

            verdict_text = ""
            if prices and all_prices:
                score, label, _, _, _ = price_score(prices[0], all_prices)
                verdict_text = "\n📊 vs market: **{}**".format(label)

            return {
                "type": "vendor_service",
                "text": (
                    "✅ **{}** quoted **{}**.\n\n{}{}\n📄 {}".format(
                        matched_vendor,
                        service,
                        price_text,
                        verdict_text,
                        ", ".join(files[:3]),
                    )
                ),
            }

        return {
            "type": "text",
            "text": "❌ **{}** has not quoted **{}**.".format(matched_vendor, service),
        }

    # --------------------------------------------------------
    # COMPARE PRICES
    # --------------------------------------------------------
    if matched_services and any(word in message for word in ["compare", "price", "cost", "expensive", "cheap", "competitive", "vs"]):
        service = matched_services[0]
        filtered = df_exploded[df_exploded["Service"] == service].drop_duplicates(subset=["Vendor", "File Name"])

        if filtered.empty:
            return {"type": "text", "text": "❌ No quotes for **{}**.".format(service)}

        vendor_prices = {}
        for _, row in filtered.iterrows():
            vendor = row["Vendor"]
            quoted_price = parse_num(str(row.get("Quoted Price", "")).strip())
            cache_key = "px_{}".format(str(row.get("File Name", "")).strip())
            cached_analysis = st.session_state.get(cache_key)
            extracted_price = cached_analysis["price_num"] if cached_analysis else 0.0
            reference_price = extracted_price if extracted_price > 0 else quoted_price

            if reference_price > 0:
                vendor_prices[vendor] = min(vendor_prices.get(vendor, reference_price), reference_price)

        if not vendor_prices:
            return {
                "type": "text",
                "text": "📋 **{}** quoted by: {}\nNo price data.".format(
                    service,
                    ", ".join(filtered["Vendor"].unique().tolist()),
                ),
            }

        avg_price = sum(vendor_prices.values()) / len(vendor_prices)
        best_vendor = min(vendor_prices, key=vendor_prices.get)
        spread = (
            round((max(vendor_prices.values()) - min(vendor_prices.values())) / min(vendor_prices.values()) * 100, 1)
            if min(vendor_prices.values()) > 0
            else 0
        )

        lines = []
        for vendor, price in sorted(vendor_prices.items(), key=lambda x: x[1]):
            pct = round((price - avg_price) / avg_price * 100, 1) if avg_price > 0 else 0
            tag = (
                "🟠 Best"
                if vendor == best_vendor
                else "⚫ Most exp." if price == max(vendor_prices.values()) else "⚪ Mid"
            )
            lines.append(
                "**{}**: {} ({:+.1f}%) {}".format(vendor, fmt_currency(price), pct, tag)
            )

        return {
            "type": "comparison",
            "text": (
                "📊 **{}**\n\n{}\n\nSpread: **{}%** · Best: **{}** at **{}**".format(
                    service,
                    "\n".join(lines),
                    spread,
                    best_vendor,
                    fmt_currency(min(vendor_prices.values())),
                )
            ),
            "service": service,
            "vendor_prices": vendor_prices,
            "avg": avg_price,
            "best_vendor": best_vendor,
        }

    # --------------------------------------------------------
    # WHO QUOTED
    # --------------------------------------------------------
    if matched_services and any(word in message for word in ["who", "vendor", "quoted", "available"]):
        service = matched_services[0]
        filtered = df_exploded[df_exploded["Service"] == service].drop_duplicates(subset=["Vendor"])

        if filtered.empty:
            return {"type": "text", "text": "❌ No vendor quoted **{}**.".format(service)}

        vendors = filtered["Vendor"].unique().tolist()
        return {
            "type": "who_quoted",
            "text": (
                "✅ **{}** vendor(s) for **{}**:\n\n{}\n\n📄 {} files".format(
                    len(vendors),
                    service,
                    "\n".join(["• **{}**".format(vendor) for vendor in vendors]),
                    df_exploded[df_exploded["Service"] == service]["File Name"].nunique(),
                )
            ),
        }

    # --------------------------------------------------------
    # VENDOR PROFILE
    # --------------------------------------------------------
    if matched_vendor:
        vendor_df = df_exploded[df_exploded["Vendor"] == matched_vendor]
        services = sorted(vendor_df["Service"].unique().tolist())
        categories = sorted(vendor_df["Category"].unique().tolist())
        quote_count = len(df_master[df_master["Vendor"] == matched_vendor])

        prices = []
        vendor_master_df = df_master[df_master["Vendor"] == matched_vendor]
        if "Quoted Price" in vendor_master_df.columns:
            for quoted in vendor_master_df["Quoted Price"]:
                value = parse_num(str(quoted))
                if value > 0:
                    prices.append(value)

        price_summary = (
            "\n\n💰 avg {} · min {} · max {}".format(
                fmt_currency(sum(prices) / len(prices)),
                fmt_currency(min(prices)),
                fmt_currency(max(prices)),
            )
            if prices
            else ""
        )

        return {
            "type": "vendor_profile",
            "text": (
                "🏢 **{}**\n\n📂 {}\n📄 {} quotes{}\n🛠 {} services:\n{}".format(
                    matched_vendor,
                    ", ".join(categories),
                    quote_count,
                    price_summary,
                    len(services),
                    "\n".join(["• {}".format(service) for service in services[:12]])
                    + ("\n…+{} more".format(len(services) - 12) if len(services) > 12 else ""),
                )
            ),
        }

    # --------------------------------------------------------
    # CATEGORY
    # --------------------------------------------------------
    if matched_category:
        category_df = df_master[df_master["Category"] == matched_category]
        prices = []

        if "Quoted Price" in category_df.columns:
            for quoted in category_df["Quoted Price"]:
                value = parse_num(str(quoted))
                if value > 0:
                    prices.append(value)

        price_text = (
            "avg {} · min {} · max {}".format(
                fmt_currency(sum(prices) / len(prices)),
                fmt_currency(min(prices)),
                fmt_currency(max(prices)),
            )
            if prices
            else "no price data"
        )

        return {
            "type": "text",
            "text": (
                "📂 **{}**\n\n📄 {} quotes · 🏢 {} vendors · 🛠 {} services\n"
                "💰 {}\n\nVendors: {}".format(
                    matched_category,
                    len(category_df),
                    category_df["Vendor"].nunique(),
                    df_exploded[df_exploded["Category"] == matched_category]["Service"].nunique(),
                    price_text,
                    ", ".join(sorted(category_df["Vendor"].unique().tolist())),
                )
            ),
        }

    # --------------------------------------------------------
    # SERVICE INFO
    # --------------------------------------------------------
    if matched_services:
        service = matched_services[0]
        filtered = df_exploded[df_exploded["Service"] == service]

        if not filtered.empty:
            vendors = filtered["Vendor"].unique().tolist()
            prices = []

            for _, row in filtered.drop_duplicates(subset=["File Name"]).iterrows():
                quoted = parse_num(str(row.get("Quoted Price", "")).strip())
                if quoted > 0:
                    prices.append(quoted)

            price_text = (
                "\n\n💰 {} – {} (avg {})".format(
                    fmt_currency(min(prices)),
                    fmt_currency(max(prices)),
                    fmt_currency(sum(prices) / len(prices)),
                )
                if prices
                else ""
            )

            return {
                "type": "service_info",
                "text": (
                    "📋 **{}**\n\n🏢 {} vendor(s): {}\n📄 {} files{}\n\n"
                    "💡 Ask *'compare {} prices'*".format(
                        service,
                        len(vendors),
                        ", ".join(vendors[:5]),
                        filtered["File Name"].nunique(),
                        price_text,
                        service,
                    )
                ),
            }

    suggestions = []
    if matched_services:
        suggestions.append("*'Compare {}?'*".format(matched_services[0]))
    if matched_vendor:
        suggestions.append("*'Profile of {}?'*".format(matched_vendor))

    fallback = "Couldn't find specific data. Try a vendor, service, or category."
    if suggestions:
        fallback += "\n\n💡 Try: " + " or ".join(suggestions)

    return {"type": "text", "text": fallback}


# ============================================================
# CHAT VISUALS
# ============================================================

def render_chat_chart(response):
    if response is None:
        return

    vendor_prices = response.get("vendor_prices") or response.get("vendor_avgs")
    avg_price = response.get("avg", 0)
    best_vendor = response.get("best_vendor", "")

    if not vendor_prices:
        return

    sorted_vendor_prices = sorted(vendor_prices.items(), key=lambda x: x[1])

    bar_colors = [
        C_ORANGE if vendor == best_vendor
        else C_DARK if price == max(vendor_prices.values())
        else C_GREY_DARK
        for vendor, price in sorted_vendor_prices
    ]

    fig = go.Figure(
        go.Bar(
            x=[vendor for vendor, _ in sorted_vendor_prices],
            y=[price for _, price in sorted_vendor_prices],
            marker_color=bar_colors,
            marker_line_width=0,
            text=[fmt_currency(price) for _, price in sorted_vendor_prices],
            textposition="outside",
            textfont=dict(size=10, color=C_DARK),
        )
    )

    if avg_price > 0:
        fig.add_hline(
            y=avg_price,
            line_dash="dash",
            line_color=C_DARK,
            line_width=1.5,
            annotation_text="Avg: {}".format(fmt_currency(avg_price)),
            annotation_position="top right",
        )

    fig.update_layout(
        height=220,
        plot_bgcolor=CBG,
        paper_bgcolor=CBG,
        margin=dict(l=5, r=10, t=16, b=5),
        font=CFONT,
        yaxis=dict(showgrid=True, gridcolor=C_GREY_LITE, zeroline=False),
        xaxis=dict(tickangle=-10, tickfont=dict(size=10)),
        bargap=0.45,
        showlegend=False,
    )

    st.plotly_chart(fig, use_container_width=True)


def build_chat_html(chat_history):
    html = "<div class='chat-outer'>"

    if not chat_history:
        html += (
            "<div class='chat-wrap'><div class='msg-bot'>"
            "👋 Hello! I have full knowledge of this catalog.<br><br>"
            "<b>Try asking:</b><br>"
            "• <i>Who quoted Cisco Catalyst?</i><br>"
            "• <i>Compare Palo Alto prices</i><br>"
            "• <i>Cheapest Cybersecurity vendor?</i><br>"
            "• <i>What does TrendMicro offer?</i><br><br>"
            "Or upload a quote file below."
            "</div></div>"
        )
    else:
        for turn in chat_history:
            html += (
                "<div class='chat-wrap'>"
                "<div class='msg-user'>{}</div>"
                "</div>".format(turn["user"])
            )

            bot_text = _chat_re.sub(
                r"\*\*(.+?)\*\*",
                r"<b>\1</b>",
                turn["bot_text"].replace("\n", "<br>"),
            )
            html += (
                "<div class='chat-wrap'>"
                "<div class='msg-bot'>{}</div>"
                "</div>".format(bot_text)
            )

    html += "</div>"
    return html


def render_quick_question_buttons(chat_key, df_master, df_exploded, chat_history_key):
    st.markdown(
        "<div style='background:#F8F8F8;border:1px solid #E0E0E0;border-top:none;padding:10px 14px 8px;margin-bottom:0'>"
        "<div style='font-size:0.70em;font-weight:700;letter-spacing:0.8px;text-transform:uppercase;color:#7D7D7D;margin-bottom:6px'>"
        "QUICK QUESTIONS</div>"
        "</div>",
        unsafe_allow_html=True,
    )

    chips = [
        "Who quoted Cisco Catalyst?",
        "Compare Palo Alto prices",
        "Cheapest Cybersecurity vendor?",
        "What does TrendMicro offer?",
        "List all vendors",
        "Catalog summary",
    ]

    row1 = st.columns(3)
    for idx, chip in enumerate(chips[:3]):
        if row1[idx].button(
            chip,
            key="chip_{}_{}".format(chat_key, chip[:20]),
            use_container_width=True,
        ):
            response = chatbot_response(chip, df_master, df_exploded)
            st.session_state[chat_history_key].append(
                {
                    "user": chip,
                    "bot_text": response["text"],
                    "bot_resp": response,
                }
            )
            st.rerun()

    row2 = st.columns(3)
    for idx, chip in enumerate(chips[3:]):
        if row2[idx].button(
            chip,
            key="chip_{}_{}".format(chat_key, chip[:20]),
            use_container_width=True,
        ):
            response = chatbot_response(chip, df_master, df_exploded)
            st.session_state[chat_history_key].append(
                {
                    "user": chip,
                    "bot_text": response["text"],
                    "bot_resp": response,
                }
            )
            st.rerun()


# ============================================================
# OVERVIEW RENDERING
# ============================================================

def render_catalog_overview(df_master, df_exploded, label="Master Catalog"):
    overview_df = df_master.copy()
    overview_df["Subcategory"] = overview_df.apply(
        lambda row: infer_subcategory(
            row.get("Category", ""),
            row.get("Comments", ""),
            row.get("File Name", ""),
        ),
        axis=1,
    )

    all_categories = sorted(
        [
            category
            for category in overview_df["Category"].unique()
            if str(category).strip() not in ["", "nan"]
        ]
    )

    k1, k2, k3, k4, k5 = st.columns(5)
    mini_kpi(k1, len(overview_df), "Total Quotations", C_ORANGE, "📄")
    mini_kpi(k2, overview_df["Vendor"].nunique(), "Unique Vendors", C_DARK, "🏢")
    mini_kpi(k3, overview_df["Category"].nunique(), "Categories", C_MID, "📂")
    mini_kpi(k4, overview_df["Subcategory"].nunique(), "Subcategories", C_GREY_DARK, "🏷️")
    mini_kpi(
        k5,
        df_exploded["Service"].nunique() if df_exploded is not None else "—",
        "Unique Services",
        C_BLACK,
        "🛠",
    )

    st.markdown("<br>", unsafe_allow_html=True)

    # --------------------------------------------------------
    # CATEGORY CARDS
    # --------------------------------------------------------
    sec("CATEGORIES AT A GLANCE")

    category_stats = []
    for category in all_categories:
        category_df = overview_df[overview_df["Category"] == category]
        service_count = (
            df_exploded[df_exploded["Category"] == category]["Service"].nunique()
            if df_exploded is not None
            else 0
        )
        category_stats.append(
            {
                "Category": category,
                "Quotations": len(category_df),
                "Vendors": category_df["Vendor"].nunique(),
                "Subcategories": category_df["Subcategory"].nunique(),
                "Services": service_count,
            }
        )

    stats_df = pd.DataFrame(category_stats).sort_values("Quotations", ascending=False)

    category_icons = {
        "Cybersecurity": "🛡️",
        "Network & Telecom": "🌐",
        "Hosting": "🖥️",
        "M365 & Power Platform": "☁️",
        "IdAM": "🔑",
        "Service Management (SNow)": "⚙️",
        "Summary & Reporting": "📊",
    }

    rows = [stats_df.iloc[i:i + 3] for i in range(0, len(stats_df), 3)]
    for chunk in rows:
        cols = st.columns(len(chunk), gap="medium")
        for col_idx, (_, row) in enumerate(chunk.iterrows()):
            category_name = row["Category"]
            icon = category_icons.get(category_name, "📁")
            category_index = all_categories.index(category_name) if category_name in all_categories else 0
            color = CHART_SEQ[category_index % len(CHART_SEQ)]

            cols[col_idx].markdown(
                "<div style='background:white;border:1px solid #E0E0E0;border-radius:6px;padding:18px 16px;border-top:4px solid {}'>"
                "<div style='font-size:1.5em;margin-bottom:6px'>{}</div>"
                "<div style='font-size:0.92em;font-weight:700;color:#2D2D2D;margin-bottom:12px;font-family:Georgia,serif'>{}</div>"
                "<div style='display:flex;gap:6px;flex-wrap:wrap'>"
                "<span style='background:#F3F3F3;border-radius:3px;padding:3px 8px;font-size:0.73em;font-weight:700;color:#2D2D2D'>📄 {} quotes</span>"
                "<span style='background:#F3F3F3;border-radius:3px;padding:3px 8px;font-size:0.73em;font-weight:700;color:#4A4A4A'>🏢 {} vendors</span>"
                "<span style='background:#F3F3F3;border-radius:3px;padding:3px 8px;font-size:0.73em;font-weight:700;color:#7D7D7D'>🛠 {} services</span>"
                "</div></div>".format(
                    color,
                    icon,
                    category_name,
                    row["Quotations"],
                    row["Vendors"],
                    row["Services"],
                ),
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)

    # --------------------------------------------------------
    # DISTRIBUTION CHARTS
    # --------------------------------------------------------
    sec("DISTRIBUTION")
    left_chart, right_chart = st.columns(2, gap="large")

    with left_chart:
        fig_quotes = go.Figure(
            go.Bar(
                x=stats_df["Category"],
                y=stats_df["Quotations"],
                marker_color=C_ORANGE,
                marker_line_width=0,
                text=stats_df["Quotations"],
                textposition="outside",
            )
        )
        pwc_bar(fig_quotes, "Quotations per Category")
        fig_quotes.update_xaxes(tickangle=-30, tickfont=dict(size=9.5))
        st.plotly_chart(fig_quotes, use_container_width=True)

    with right_chart:
        fig_vendors = go.Figure(
            go.Bar(
                x=stats_df["Category"],
                y=stats_df["Vendors"],
                marker_color=C_DARK,
                marker_line_width=0,
                text=stats_df["Vendors"],
                textposition="outside",
            )
        )
        pwc_bar(fig_vendors, "Vendors per Category")
        fig_vendors.update_xaxes(tickangle=-30, tickfont=dict(size=9.5))
        st.plotly_chart(fig_vendors, use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # --------------------------------------------------------
    # CATALOG COMPOSITION
    # --------------------------------------------------------
    sec("CATALOG COMPOSITION")
    left_panel, right_panel = st.columns([1, 2], gap="large")

    with left_panel:
        fig_pie = px.pie(
            stats_df,
            values="Quotations",
            names="Category",
            hole=0.55,
            color_discrete_sequence=CHART_SEQ,
        )
        fig_pie.update_traces(
            textposition="outside",
            textinfo="percent+label",
            textfont_size=10,
            pull=[0.03] * len(stats_df),
        )
        fig_pie.update_layout(
            height=360,
            margin=dict(l=10, r=10, t=10, b=10),
            paper_bgcolor=CBG,
            font=CFONT,
            showlegend=False,
        )
        st.plotly_chart(fig_pie, use_container_width=True)

    with right_panel:
        table_rows = [
            "<table class='comp-table'><thead><tr>"
            "<th>Category</th>"
            "<th style='text-align:center'>Quotes</th>"
            "<th style='text-align:center'>Vendors</th>"
            "<th style='text-align:center'>Services</th>"
            "<th style='text-align:center'>Subcats</th>"
            "</tr></thead><tbody>"
        ]

        for _, row in stats_df.iterrows():
            category_name = row["Category"]
            category_index = all_categories.index(category_name) if category_name in all_categories else 0
            color = CHART_SEQ[category_index % len(CHART_SEQ)]
            icon = category_icons.get(category_name, "📁")

            table_rows.append(
                "<tr><td><span style='border-left:4px solid {};padding-left:8px;font-weight:600'>"
                "{} {}</span></td>"
                "<td style='text-align:center;font-weight:700;color:#D04A02'>{}</td>"
                "<td style='text-align:center;font-weight:700;color:#4A4A4A'>{}</td>"
                "<td style='text-align:center;font-weight:700;color:#7D7D7D'>{}</td>"
                "<td style='text-align:center;font-weight:700;color:#2D2D2D'>{}</td>"
                "</tr>".format(
                    color,
                    icon,
                    category_name,
                    row["Quotations"],
                    row["Vendors"],
                    row["Services"],
                    row["Subcategories"],
                )
            )

        table_rows.append("</tbody></table>")
        render_html_table(table_rows)

    # --------------------------------------------------------
    # DRILL-DOWN BY CATEGORY
    # --------------------------------------------------------
    st.markdown("<br>", unsafe_allow_html=True)
    sec("DRILL-DOWN BY CATEGORY")

    ordered_categories = ["Cybersecurity"] + [c for c in all_categories if c != "Cybersecurity"]
    tabs = st.tabs(
        [
            "🛡️ {}".format(category) if category == "Cybersecurity" else category
            for category in ordered_categories
        ]
    )

    for tab_idx, category_name in enumerate(ordered_categories):
        with tabs[tab_idx]:
            category_df = overview_df[overview_df["Category"] == category_name].copy()
            if category_df.empty:
                st.info("No data.")
                continue

            c1, c2, c3, c4 = st.columns(4)
            service_count = (
                df_exploded[df_exploded["Category"] == category_name]["Service"].nunique()
                if df_exploded is not None
                else 0
            )

            mini_kpi(c1, len(category_df), "Quotations", C_ORANGE, "📄")
            mini_kpi(c2, category_df["Vendor"].nunique(), "Vendors", C_DARK, "🏢")
            mini_kpi(c3, service_count, "Services", C_MID, "🛠")
            mini_kpi(c4, category_df["Subcategory"].nunique(), "Subcats", C_GREY_DARK, "🏷️")

            st.markdown("<br>", unsafe_allow_html=True)

            subcategory_stats = (
                category_df.groupby("Subcategory")
                .agg(
                    Quotations=("File Name", "count"),
                    Vendors=("Vendor", "nunique"),
                )
                .reset_index()
                .sort_values("Quotations", ascending=False)
            )

            left_sub, right_sub = st.columns([1, 1], gap="medium")

            with left_sub:
                st.markdown(
                    "<div style='font-size:0.77em;font-weight:700;text-transform:uppercase;color:#2D2D2D;margin-bottom:8px'>Subcategories</div>",
                    unsafe_allow_html=True,
                )
                fig_sub = go.Figure(
                    go.Bar(
                        x=subcategory_stats["Quotations"],
                        y=subcategory_stats["Subcategory"],
                        orientation="h",
                        marker_color=C_ORANGE,
                        marker_line_width=0,
                        text=subcategory_stats["Quotations"],
                        textposition="outside",
                    )
                )
                fig_sub.update_layout(
                    height=max(200, len(subcategory_stats) * 36),
                    plot_bgcolor=CBG,
                    paper_bgcolor=CBG,
                    margin=dict(l=5, r=40, t=8, b=8),
                    font=CFONT,
                    xaxis=dict(showgrid=True, gridcolor=C_GREY_LITE, zeroline=False),
                    yaxis=dict(autorange="reversed", tickfont=dict(size=9.5)),
                    bargap=0.3,
                )
                st.plotly_chart(fig_sub, use_container_width=True)

            with right_sub:
                st.markdown(
                    "<div style='font-size:0.77em;font-weight:700;text-transform:uppercase;color:#2D2D2D;margin-bottom:8px'>Vendors</div>",
                    unsafe_allow_html=True,
                )
                vendor_stats = (
                    category_df.groupby("Vendor")
                    .agg(Quotations=("File Name", "count"))
                    .reset_index()
                    .sort_values("Quotations", ascending=False)
                )
                fig_vendor = go.Figure(
                    go.Bar(
                        x=vendor_stats["Quotations"],
                        y=vendor_stats["Vendor"],
                        orientation="h",
                        marker_color=C_DARK,
                        marker_line_width=0,
                        text=vendor_stats["Quotations"],
                        textposition="outside",
                    )
                )
                fig_vendor.update_layout(
                    height=max(200, len(vendor_stats) * 36),
                    plot_bgcolor=CBG,
                    paper_bgcolor=CBG,
                    margin=dict(l=5, r=40, t=8, b=8),
                    font=CFONT,
                    xaxis=dict(showgrid=True, gridcolor=C_GREY_LITE, zeroline=False),
                    yaxis=dict(autorange="reversed", tickfont=dict(size=9.5)),
                    bargap=0.3,
                )
                st.plotly_chart(fig_vendor, use_container_width=True)

            table_rows = [
                "<table class='comp-table'><thead><tr>"
                "<th>File Name</th><th>Vendor</th><th>Subcategory</th><th>Services</th>"
                "</tr></thead><tbody>"
            ]

            vendor_map = st.session_state.get("vendor_color_map", {})

            for idx, (_, row) in enumerate(category_df.sort_values("Subcategory").iterrows()):
                bg = "white" if idx % 2 == 0 else "#F8F8F8"
                vendor_color = vendor_map.get(row["Vendor"], C_DARK)
                comments = str(row.get("Comments", "")).replace("\n", " · ")[:80]

                table_rows.append(
                    "<tr style='background:{}'>"
                    "<td style='font-family:monospace;font-size:0.77em;word-break:break-all'>{}</td>"
                    "<td>{}</td>"
                    "<td style='color:#7D7D7D;font-size:0.81em'>{}</td>"
                    "<td style='font-size:0.79em'>{}</td>"
                    "</tr>".format(
                        bg,
                        row.get("File Name", ""),
                        vpill(row["Vendor"], vendor_color),
                        row["Subcategory"],
                        comments,
                    )
                )

            table_rows.append("</tbody></table>")
            render_html_table(table_rows)

# ============================================================
# PART 3 — BROWSE & VERDICT
# ============================================================

def init_chat_state(chat_key: str):
    chat_history_key = f"chat_history_{chat_key}"
    if chat_history_key not in st.session_state:
        st.session_state[chat_history_key] = []
    return chat_history_key


def append_chat_message(chat_history_key: str, user_text: str, response: dict):
    st.session_state[chat_history_key].append(
        {
            "user": user_text,
            "bot_text": response["text"],
            "bot_resp": response,
        }
    )


def render_chat_panel(df_master, df_exploded, chat_key: str):
    chat_history_key = init_chat_state(chat_key)

    st.markdown(
        "<div class='chat-header'>"
        "<span style='font-weight:700;font-size:0.96em;font-family:Georgia,serif'>"
        "💬 Procurement Assistant</span>"
        "<span style='font-size:0.74em;opacity:0.6;margin-left:10px'>"
        "Full catalog knowledge · Upload a quote</span>"
        "</div>",
        unsafe_allow_html=True,
    )

    st.markdown(
        build_chat_html(st.session_state[chat_history_key]),
        unsafe_allow_html=True,
    )

    render_quick_question_buttons(chat_key, df_master, df_exploded, chat_history_key)

    if st.session_state[chat_history_key]:
        last_response = st.session_state[chat_history_key][-1].get("bot_resp")
        if last_response:
            render_chat_chart(last_response)

    uploaded_chat_file = st.file_uploader(
        "📎 Upload quote file for instant scoring",
        type=["pdf", "xlsx", "xls", "docx"],
        key=f"chat_file_{chat_key}",
    )

    with st.form(f"chat_form_{chat_key}", clear_on_submit=True):
        c1, c2 = st.columns([5, 1])
        with c1:
            user_input = st.text_input(
                "msg",
                placeholder="Ask a question…",
                label_visibility="collapsed",
            )
        with c2:
            sent = st.form_submit_button(
                "Send",
                type="primary",
                use_container_width=True,
            )

    if sent and user_input.strip():
        response = chatbot_response(user_input.strip(), df_master, df_exploded)
        append_chat_message(chat_history_key, user_input.strip(), response)
        st.rerun()

    last_file_key = f"last_chat_file_{chat_key}"
    if (
        uploaded_chat_file is not None
        and st.session_state.get(last_file_key) != uploaded_chat_file.name
    ):
        st.session_state[last_file_key] = uploaded_chat_file.name
        file_bytes = uploaded_chat_file.read()
        response = chatbot_response(
            "uploaded file",
            df_master,
            df_exploded,
            uploaded_file_bytes=file_bytes,
            uploaded_file_name=uploaded_chat_file.name,
        )
        append_chat_message(chat_history_key, f"📎 {uploaded_chat_file.name}", response)
        st.rerun()

    if st.button("🗑 Clear chat", key=f"clr_{chat_key}"):
        st.session_state[chat_history_key] = []
        st.rerun()


def vendor_service_matrix(selected_df: pd.DataFrame, selected_services: list[str]) -> pd.DataFrame:
    heat = []
    vendors = sorted(selected_df["Vendor"].unique())

    for service in selected_services:
        for vendor in vendors:
            has_quote = int(
                len(
                    selected_df[
                        (selected_df["Service"] == service)
                        & (selected_df["Vendor"] == vendor)
                    ]
                ) > 0
            )
            heat.append(
                {
                    "Service": service[:40],
                    "Vendor": vendor,
                    "Covered": has_quote,
                }
            )
    return pd.DataFrame(heat)


def extract_vendor_price_map(selected_df: pd.DataFrame, chat_key: str) -> dict:
    vendor_prices = {}

    deduped = selected_df.drop_duplicates(subset=["Vendor", "File Name"])
    for _, row in deduped.iterrows():
        vendor = row["Vendor"]
        quoted_price = parse_num(str(row.get("Quoted Price", "")).strip())

        cache_key = "px_{}_{}".format(chat_key, str(row.get("File Name", "")).strip())
        cached_analysis = st.session_state.get(cache_key)
        extracted_price = cached_analysis["price_num"] if cached_analysis else 0.0

        reference_price = extracted_price if extracted_price > 0 else quoted_price
        if reference_price > 0:
            vendor_prices[vendor] = min(vendor_prices.get(vendor, reference_price), reference_price)

    if vendor_prices:
        return vendor_prices

    for _, row in selected_df.drop_duplicates(subset=["Vendor"]).iterrows():
        vendor = row["Vendor"]
        quoted_price = parse_num(str(row.get("Quoted Price", "")).strip())
        if quoted_price > 0 and vendor not in vendor_prices:
            vendor_prices[vendor] = quoted_price

    return vendor_prices


def render_service_competitiveness_map(df_exploded: pd.DataFrame):
    sec(
        "SERVICE COMPETITIVENESS MAP",
        "Orange = multiple vendors · Grey = single vendor only",
    )

    service_summary = (
        df_exploded.groupby("Service")["Vendor"]
        .nunique()
        .reset_index()
        .sort_values("Vendor", ascending=False)
    )
    service_summary.columns = ["Service", "Vendor Count"]

    top_20 = service_summary.head(20)

    fig = go.Figure(
        go.Bar(
            x=top_20["Vendor Count"],
            y=top_20["Service"].apply(lambda x: x[:48]),
            orientation="h",
            marker_color=[
                C_ORANGE if count > 1 else C_GREY
                for count in top_20["Vendor Count"]
            ],
            marker_line_width=0,
            text=top_20["Vendor Count"],
            textposition="outside",
            textfont=dict(size=10),
        )
    )

    fig.update_layout(
        height=520,
        plot_bgcolor=CBG,
        paper_bgcolor=CBG,
        margin=dict(l=5, r=40, t=10, b=8),
        font=CFONT,
        xaxis=dict(
            title="Number of Vendors",
            showgrid=True,
            gridcolor=C_GREY_LITE,
            zeroline=False,
        ),
        yaxis=dict(autorange="reversed", tickfont=dict(size=9.2)),
        bargap=0.28,
        showlegend=False,
    )
    st.plotly_chart(fig, use_container_width=True)

    multi_vendor = service_summary[service_summary["Vendor Count"] > 1].shape[0]
    single_vendor = service_summary[service_summary["Vendor Count"] == 1].shape[0]

    k1, k2, k3 = st.columns(3)
    kpi_box(k1, len(service_summary), "Total Services", C_DARK)
    kpi_box(k2, multi_vendor, "Competitive (2+ vendors)", C_ORANGE)
    kpi_box(k3, single_vendor, "Single Vendor Only", C_GREY_DARK)


def render_coverage_verdict(selected_df: pd.DataFrame, selected_services: list[str]):
    vendor_service_map = defaultdict(set)
    for _, row in selected_df.iterrows():
        vendor_service_map[row["Vendor"]].add(row["Service"])

    full_coverage = [
        vendor for vendor, services in vendor_service_map.items()
        if set(selected_services).issubset(services)
    ]
    partial_coverage = [
        vendor for vendor, services in vendor_service_map.items()
        if set(selected_services) & services and vendor not in full_coverage
    ]

    if len(selected_services) == 1:
        vendor_count = selected_df["Vendor"].nunique()
        if vendor_count > 1:
            st.markdown(
                "<div class='verdict-good'>"
                "✅ <b>{}</b> — quoted by <b>{} vendors</b>."
                "</div>".format(selected_services[0], vendor_count),
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                "<div class='verdict-mid'>"
                "⚠️ <b>{}</b> — only <b>1 vendor</b>."
                "</div>".format(selected_services[0]),
                unsafe_allow_html=True,
            )
    else:
        if full_coverage:
            st.markdown(
                "<div class='verdict-good'>"
                "✅ <b>{}</b> vendor(s) cover ALL {}: <b>{}</b>"
                "</div>".format(
                    len(full_coverage),
                    len(selected_services),
                    ", ".join(full_coverage),
                ),
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                "<div class='verdict-mid'>"
                "⚠️ No single vendor covers all {}. Partial: <b>{}</b>"
                "</div>".format(
                    len(selected_services),
                    ", ".join(partial_coverage) if partial_coverage else "None",
                ),
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)


def render_vendor_service_heatmap(selected_df: pd.DataFrame, selected_services: list[str]):
    sec("VENDOR × SERVICE COVERAGE MAP", "✅ = vendor has quoted this service")

    heat_df = vendor_service_matrix(selected_df, selected_services)
    if heat_df.empty:
        return

    pivot = heat_df.pivot_table(
        index="Service",
        columns="Vendor",
        values="Covered",
        fill_value=0,
    )

    fig = go.Figure(
        go.Heatmap(
            z=pivot.values,
            x=pivot.columns.tolist(),
            y=pivot.index.tolist(),
            colorscale=[[0, C_GREY_LITE], [1, C_ORANGE]],
            showscale=False,
            text=[
                ["✅" if val == 1 else "—" for val in row]
                for row in pivot.values
            ],
            texttemplate="%{text}",
            textfont=dict(size=16),
        )
    )

    fig.update_layout(
        height=max(180, len(selected_services) * 65),
        plot_bgcolor=CBG,
        paper_bgcolor=CBG,
        margin=dict(l=5, r=10, t=10, b=10),
        font=CFONT,
        xaxis=dict(tickangle=-20, tickfont=dict(size=10)),
        yaxis=dict(tickfont=dict(size=9.5), autorange="reversed"),
    )

    st.plotly_chart(fig, use_container_width=True)


def render_quotes_per_vendor_chart(selected_df: pd.DataFrame, vcmap: dict):
    sec("QUOTES PER VENDOR", "Number of quote files per vendor")

    vendor_quote_df = (
        selected_df.drop_duplicates(subset=["Vendor", "File Name"])
        .groupby("Vendor")
        .size()
        .reset_index()
    )
    vendor_quote_df.columns = ["Vendor", "Quotes"]
    vendor_quote_df = vendor_quote_df.sort_values("Quotes", ascending=False)

    fig = go.Figure(
        go.Bar(
            x=vendor_quote_df["Vendor"],
            y=vendor_quote_df["Quotes"],
            marker_color=[vcmap.get(vendor, C_DARK) for vendor in vendor_quote_df["Vendor"]],
            marker_line_width=0,
            text=vendor_quote_df["Quotes"],
            textposition="outside",
        )
    )

    fig.update_layout(
        height=260,
        plot_bgcolor=CBG,
        paper_bgcolor=CBG,
        margin=dict(l=5, r=10, t=10, b=8),
        font=CFONT,
        yaxis=dict(title="Quote Files", showgrid=True, gridcolor=C_GREY_LITE, zeroline=False),
        xaxis=dict(tickangle=-15, tickfont=dict(size=10.5)),
        bargap=0.4,
        showlegend=False,
    )
    st.plotly_chart(fig, use_container_width=True)


def render_vendors_per_service_chart(selected_df: pd.DataFrame, selected_services: list[str]):
    if len(selected_services) <= 1:
        return

    sec("VENDORS PER SERVICE")

    rows = []
    for service in selected_services:
        vendor_count = selected_df[selected_df["Service"] == service]["Vendor"].nunique()
        rows.append({"Service": service[:45], "Vendors": vendor_count})

    service_vendor_df = pd.DataFrame(rows)

    fig = go.Figure(
        go.Bar(
            x=service_vendor_df["Service"],
            y=service_vendor_df["Vendors"],
            marker_color=[
                C_ORANGE if count > 1 else C_GREY_DARK
                for count in service_vendor_df["Vendors"]
            ],
            marker_line_width=0,
            text=service_vendor_df["Vendors"],
            textposition="outside",
        )
    )

    fig.update_layout(
        height=260,
        plot_bgcolor=CBG,
        paper_bgcolor=CBG,
        margin=dict(l=5, r=10, t=10, b=8),
        font=CFONT,
        yaxis=dict(title="Vendors", showgrid=True, gridcolor=C_GREY_LITE, zeroline=False),
        xaxis=dict(tickangle=-15, tickfont=dict(size=10)),
        bargap=0.4,
        showlegend=False,
    )
    st.plotly_chart(fig, use_container_width=True)


def render_price_verdict_cards(vendor_prices: dict):
    avg_price = sum(vendor_prices.values()) / len(vendor_prices)
    best_vendor = min(vendor_prices, key=vendor_prices.get)
    spread = (
        round((max(vendor_prices.values()) - min(vendor_prices.values())) / min(vendor_prices.values()) * 100, 1)
        if min(vendor_prices.values()) > 0
        else 0
    )

    sec("PRICE VERDICT")

    c1, c2, c3 = st.columns(3)

    c1.markdown(
        "<div class='scard scard-orange'>"
        "<div style='font-size:0.67em;font-weight:700;text-transform:uppercase;color:#D04A02'>"
        "Best Price</div>"
        "<div style='font-size:1.5em;font-weight:800;color:#D04A02;font-family:Georgia,serif'>"
        "{}</div>"
        "<div style='font-size:0.73em;color:#7D7D7D;margin-top:2px'>{}</div>"
        "</div>".format(
            fmt_currency(min(vendor_prices.values())),
            best_vendor,
        ),
        unsafe_allow_html=True,
    )

    c2.markdown(
        "<div class='scard scard-dark'>"
        "<div style='font-size:0.67em;font-weight:700;text-transform:uppercase;color:#2D2D2D'>"
        "Market Average</div>"
        "<div style='font-size:1.5em;font-weight:800;color:#2D2D2D;font-family:Georgia,serif'>"
        "{}</div>"
        "<div style='font-size:0.73em;color:#7D7D7D;margin-top:2px'>{} vendors</div>"
        "</div>".format(
            fmt_currency(avg_price),
            len(vendor_prices),
        ),
        unsafe_allow_html=True,
    )

    c3.markdown(
        "<div class='scard scard-grey'>"
        "<div style='font-size:0.67em;font-weight:700;text-transform:uppercase;color:#7D7D7D'>"
        "Price Spread</div>"
        "<div style='font-size:1.5em;font-weight:800;color:#4A4A4A;font-family:Georgia,serif'>"
        "{}%</div>"
        "<div style='font-size:0.73em;color:#7D7D7D;margin-top:2px'>negotiation room</div>"
        "</div>".format(spread),
        unsafe_allow_html=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)
    return avg_price, best_vendor, spread


def render_price_comparison_chart(vendor_prices: dict, avg_price: float, best_vendor: str):
    sec("PRICE COMPARISON CHART", "Orange = best · Dark = highest · Dashed = market average")

    sorted_vendor_prices = sorted(vendor_prices.items(), key=lambda x: x[1])
    worst_vendor = max(vendor_prices, key=vendor_prices.get)

    colors = [
        C_ORANGE if vendor == best_vendor
        else C_DARK if vendor == worst_vendor
        else C_GREY_DARK
        for vendor, _ in sorted_vendor_prices
    ]

    fig = go.Figure(
        go.Bar(
            x=[vendor for vendor, _ in sorted_vendor_prices],
            y=[price for _, price in sorted_vendor_prices],
            marker_color=colors,
            marker_line_width=0,
            text=[fmt_currency(price) for _, price in sorted_vendor_prices],
            textposition="outside",
            textfont=dict(size=11, color=C_DARK),
        )
    )

    fig.add_hline(
        y=avg_price,
        line_dash="dash",
        line_color=C_MID,
        line_width=2,
        annotation_text="Avg: {}".format(fmt_currency(avg_price)),
        annotation_position="top right",
    )

    fig.update_layout(
        height=290,
        plot_bgcolor=CBG,
        paper_bgcolor=CBG,
        margin=dict(l=5, r=10, t=24, b=8),
        font=CFONT,
        yaxis=dict(title="Price (USD)", showgrid=True, gridcolor=C_GREY_LITE, zeroline=False),
        xaxis=dict(tickangle=-10, tickfont=dict(size=10.5)),
        bargap=0.4,
        showlegend=False,
    )
    st.plotly_chart(fig, use_container_width=True)


def render_vendor_scorecard(vendor_prices: dict, vcmap: dict, avg_price: float, spread: float):
    sec("VENDOR SCORE CARD")

    all_prices = list(vendor_prices.values())

    table_rows = [
        "<table class='comp-table'><thead><tr>"
        "<th>Rank</th><th>Vendor</th><th>Price</th><th>vs Average</th><th>Score</th><th>Verdict</th>"
        "</tr></thead><tbody>"
    ]

    for rank, (vendor, price) in enumerate(sorted(vendor_prices.items(), key=lambda x: x[1]), 1):
        bg = "white" if rank % 2 == 0 else "#F8F8F8"
        vendor_color = vcmap.get(vendor, C_DARK)
        other_prices = [p for p in all_prices if p != price]

        score = None
        if price > 0 and other_prices:
            score, _, _, _, _ = price_score(price, other_prices)

        score_col = score_color(score)
        pct = round((price - avg_price) / avg_price * 100, 1) if avg_price > 0 else 0

        vs_text = (
            "{}% below ✅".format(abs(pct))
            if pct < 0
            else "{}% above ⚠️".format(abs(pct)) if pct > 0 else "At average"
        )
        vs_color = C_ORANGE if pct < 0 else C_DARK if pct > 10 else C_GREY_DARK

        verdict_text, _, verdict_color = get_verdict(score)

        medal = "🥇" if rank == 1 else "🥈" if rank == 2 else "🥉" if rank == 3 else str(rank)

        table_rows.append(
            "<tr style='background:{}'>"
            "<td style='text-align:center;font-size:1.05em'>{}</td>"
            "<td>{}</td>"
            "<td style='font-family:monospace;font-weight:700'>{}</td>"
            "<td style='color:{}'>{}</td>"
            "<td style='text-align:center'><span style='font-weight:800;font-size:1.1em;color:{}'>{}</span></td>"
            "<td style='font-weight:700;color:{}'>{}</td>"
            "</tr>".format(
                bg,
                medal,
                vpill(vendor, vendor_color),
                fmt_currency(price),
                vs_color,
                vs_text,
                score_col,
                score if score is not None else "—",
                verdict_color,
                verdict_text,
            )
        )

    table_rows.append("</tbody></table>")
    render_html_table(table_rows)

    pct_below = round((avg_price - min(vendor_prices.values())) / avg_price * 100, 1) if avg_price > 0 else 0
    insight(
        "<b>{}</b> is most competitive at <b>{}</b> — <b>{}% below</b> market avg. "
        "Spread <b>{}%</b> → <b>{}</b>.".format(
            min(vendor_prices, key=vendor_prices.get),
            fmt_currency(min(vendor_prices.values())),
            pct_below,
            spread,
            "strong negotiation potential" if spread > 20 else "moderate room" if spread > 10 else "competitive market",
        )
    )


def extract_prices_for_service_rows(service_df: pd.DataFrame, chat_key: str):
    progress = st.progress(0)
    total = len(service_df)

    for idx, (_, row) in enumerate(service_df.iterrows()):
        filename = str(row.get("File Name", "")).strip()
        cache_key = f"px_{chat_key}_{filename}"

        if not st.session_state.get(cache_key):
            local_path = os.path.join(DEMO_DIR, filename)

            if os.path.exists(local_path):
                st.session_state[cache_key] = extract_price_from_file(local_path)
            else:
                url = resolve_url(row)
                if url and url.startswith("http") and REQUESTS_OK:
                    try:
                        response = requests.get(url, timeout=20)
                        ext = url.split("?")[0].rsplit(".", 1)[-1].lower()
                        st.session_state[cache_key] = extract_price_from_bytes(response.content, ext)
                    except Exception:
                        pass

        progress.progress((idx + 1) / total)

    progress.empty()


def render_quote_file_details(selected_df: pd.DataFrame, selected_services: list[str], vcmap: dict, has_prices: bool, chat_key: str):
    st.markdown("<br>", unsafe_allow_html=True)
    sec("QUOTATION FILE DETAILS")

    has_price_column = "Quoted Price" in selected_df.columns

    for service in selected_services:
        service_df = (
            selected_df[selected_df["Service"] == service]
            .drop_duplicates(subset=["Vendor", "File Name"])
            .sort_values("Vendor")
        )

        vendor_count = service_df["Vendor"].nunique()

        st.markdown(
            "<div style='background:white;border-left:4px solid {};padding:10px 14px;border-radius:2px;margin:8px 0;font-weight:700;font-size:0.88em'>"
            "{}  ·  {} vendor(s)  ·  {}"
            "</div>".format(
                C_ORANGE if vendor_count > 1 else C_GREY_DARK,
                service,
                vendor_count,
                "✅ COMPETITIVE" if vendor_count > 1 else "⚠️ SINGLE VENDOR",
            ),
            unsafe_allow_html=True,
        )

        all_prices = []
        for _, row in service_df.iterrows():
            quoted_price = parse_num(str(row.get("Quoted Price", "")).strip())
            if quoted_price > 0:
                all_prices.append(quoted_price)

        table_rows = [
            "<table class='comp-table'><thead><tr>"
            "<th>Vendor</th><th>File</th>"
        ]
        if has_price_column:
            table_rows.append("<th>Quoted Price</th>")
        table_rows.append("<th>Score</th><th>Verdict</th><th>Open</th></tr></thead><tbody>")

        for idx, (_, row) in enumerate(service_df.iterrows()):
            bg = "white" if idx % 2 == 0 else "#F8F8F8"
            vendor_color = vcmap.get(row["Vendor"], C_DARK)
            filename = str(row.get("File Name", "")).strip()
            url = resolve_url(row)

            quoted_price_num = parse_num(str(row.get("Quoted Price", "")).strip())
            cache_key = f"px_{chat_key}_{filename}"
            cached_analysis = st.session_state.get(cache_key)

            reference_price = (
                cached_analysis["price_num"]
                if cached_analysis and cached_analysis.get("price_num", 0) > 0
                else quoted_price_num if quoted_price_num > 0 else 0
            )

            other_prices = [p for p in all_prices if p != reference_price]
            score = None
            verdict = "—"
            verdict_color = C_GREY_DARK

            if reference_price > 0 and other_prices:
                score, _, _, _, _ = price_score(reference_price, other_prices)
                verdict, _, verdict_color = get_verdict(score)

            score_col = score_color(score)

            link = (
                "<a href='{}' target='_blank' style='color:#D04A02;font-weight:600;text-decoration:none'>📂 Open</a>".format(url)
                if url
                else "—"
            )

            row_html = (
                "<tr style='background:{}'>"
                "<td>{}</td>"
                "<td style='font-family:monospace;font-size:0.77em;word-break:break-all'>{}</td>".format(
                    bg,
                    vpill(row["Vendor"], vendor_color),
                    filename,
                )
            )

            if has_price_column:
                row_html += (
                    "<td style='font-family:monospace;font-weight:700;color:#D04A02'>{}</td>".format(
                        fmt_currency(quoted_price_num) if quoted_price_num > 0 else "—"
                    )
                )

            row_html += (
                "<td style='text-align:center'><span style='font-weight:800;color:{}'>{}</span></td>"
                "<td style='font-weight:700;color:{}'>{}</td>"
                "<td>{}</td>"
                "</tr>".format(
                    score_col,
                    "{}/100".format(score) if score is not None else "—",
                    verdict_color,
                    verdict,
                    link,
                )
            )

            table_rows.append(row_html)

        table_rows.append("</tbody></table>")
        render_html_table(table_rows)

        if has_prices or os.path.exists(DEMO_DIR):
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button(
                "🔍 Extract Prices — {}".format(service[:38]),
                key="ep_{}_{}".format(chat_key, service[:28]),
                type="primary",
            ):
                extract_prices_for_service_rows(service_df, chat_key)
                st.rerun()


def render_browse_verdict(
    df_master,
    df_exploded,
    vcmap,
    label="",
    has_prices=False,
    chat_key_suffix="",
):
    st.session_state["vendor_color_map"] = vcmap

    st.markdown("<div class='filter-bar'>", unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3)

    all_categories = ["All"] + sorted(
        [
            category
            for category in df_master["Category"].unique()
            if str(category).strip() not in ["", "nan"]
        ]
    )

    with c1:
        st.markdown(
            "<p style='color:#D04A02;font-size:0.77em;font-weight:700;margin-bottom:4px;letter-spacing:0.6px;text-transform:uppercase'>📂 CATEGORY</p>",
            unsafe_allow_html=True,
        )
        selected_category = st.selectbox(
            "cat",
            all_categories,
            label_visibility="collapsed",
            key=f"bv_cat_{chat_key_suffix}",
        )

    vendor_pool_df = df_master if selected_category == "All" else df_master[df_master["Category"] == selected_category]
    all_vendors = ["All"] + sorted(
        [
            vendor
            for vendor in vendor_pool_df["Vendor"].unique()
            if str(vendor).strip() not in ["", "nan"]
        ]
    )

    with c2:
        st.markdown(
            "<p style='color:#D04A02;font-size:0.77em;font-weight:700;margin-bottom:4px;letter-spacing:0.6px;text-transform:uppercase'>🏢 VENDOR</p>",
            unsafe_allow_html=True,
        )
        selected_vendor = st.selectbox(
            "ven",
            all_vendors,
            label_visibility="collapsed",
            key=f"bv_ven_{chat_key_suffix}",
        )

    with c3:
        st.markdown(
            "<p style='color:#D04A02;font-size:0.77em;font-weight:700;margin-bottom:4px;letter-spacing:0.6px;text-transform:uppercase'>🔍 SEARCH SERVICE</p>",
            unsafe_allow_html=True,
        )
        service_filter = st.text_input(
            "svc",
            placeholder="Type to filter…",
            label_visibility="collapsed",
            key=f"bv_svc_{chat_key_suffix}",
        )

    st.markdown("</div>", unsafe_allow_html=True)

    filtered_df = df_exploded.copy()
    if selected_category != "All":
        filtered_df = filtered_df[filtered_df["Category"] == selected_category]
    if selected_vendor != "All":
        filtered_df = filtered_df[filtered_df["Vendor"] == selected_vendor]

    available_services = sorted(
        [
            service
            for service in filtered_df["Service"].unique()
            if str(service).strip() not in ["", "nan"]
        ]
    )

    if service_filter:
        available_services = [
            service for service in available_services
            if service_filter.lower() in service.lower()
        ]

    left_col, right_col = st.columns([1, 1], gap="large")

    with left_col:
        render_chat_panel(df_master, df_exploded, chat_key_suffix)

    with right_col:
        sec("SERVICE BROWSER")

        selected_services = st.multiselect(
            "Select services to analyse",
            options=available_services,
            default=[],
            label_visibility="visible",
            key=f"sel_svcs_{chat_key_suffix}",
        )

        if not selected_services:
            render_service_competitiveness_map(filtered_df)
            return

        selected_df = filtered_df[filtered_df["Service"].isin(selected_services)].copy()
        if selected_df.empty:
            st.warning("No quotations found.")
            return

        vendor_prices = extract_vendor_price_map(selected_df, chat_key_suffix)

        render_coverage_verdict(selected_df, selected_services)
        render_vendor_service_heatmap(selected_df, selected_services)
        render_quotes_per_vendor_chart(selected_df, vcmap)
        render_vendors_per_service_chart(selected_df, selected_services)

        if vendor_prices:
            avg_price, best_vendor, spread = render_price_verdict_cards(vendor_prices)
            render_price_comparison_chart(vendor_prices, avg_price, best_vendor)
            render_vendor_scorecard(vendor_prices, vcmap, avg_price, spread)
        elif not has_prices:
            st.markdown(
                "<div class='insight-box'>"
                "ℹ️ Price data is not available for the Master Catalog — quotation files are stored on "
                "SharePoint. Switch to the <b>Dummy Data</b> tabs to see full price analysis."
                "</div>",
                unsafe_allow_html=True,
            )

        render_quote_file_details(
            selected_df=selected_df,
            selected_services=selected_services,
            vcmap=vcmap,
            has_prices=has_prices,
            chat_key=chat_key_suffix,
        )

# ============================================================
# PART 4 — APP STATE, DATA LOADING, MAIN LAYOUT, TABS
# ============================================================

def init_session_defaults():
    defaults = [
        ("tab2_upload_price", 0.0),
        ("tab2_upload_fname", ""),
        ("tab2_file_bytes", None),
        ("tab2_file_ext", ""),
        ("chat_redirect_upload", False),
        ("gh_prices_loaded", False),
        ("real_analysis_done", False),
        ("real_analysis_df", None),
    ]

    for key, value in defaults:
        if key not in st.session_state:
            st.session_state[key] = value


def apply_uploaded_catalog_if_present(df_master, df_exp_master):
    if "uploaded_catalog_df" in st.session_state and "uploaded_catalog_exp" in st.session_state:
        return st.session_state["uploaded_catalog_df"], st.session_state["uploaded_catalog_exp"]
    return df_master, df_exp_master


def render_main_header():
    st.markdown(
        "<div style='background:#2D2D2D;color:white;padding:24px 32px;border-radius:4px;"
        "border-left:8px solid #D04A02;margin-bottom:20px'>"
        "<div style='font-size:0.68em;font-weight:700;letter-spacing:2.5px;text-transform:uppercase;"
        "color:#D04A02;margin-bottom:8px'>"
        "IT PROCUREMENT · INTELLIGENCE DASHBOARD</div>"
        "<h1 style='margin:0;font-size:1.85em;font-weight:700;color:white;font-family:Georgia,serif'>"
        "Procurement Intelligence Dashboard</h1>"
        "<p style='margin:8px 0 0;opacity:0.5;font-size:0.85em'>"
        "Master Catalog (Excel) · Dummy Data (CSV) · Browse &amp; Verdict · Upload &amp; Score · "
        "Vendor Analysis</p>"
        "</div>",
        unsafe_allow_html=True,
    )


def render_top_kpis(df_master, df_exp_master, df_dummy, df_exp_dummy, no_master, no_dummy):
    left, right = st.columns(2, gap="large")

    with left:
        st.markdown(
            "<div class='bucket-header'>📁 MASTER CATALOG (Excel)</div>",
            unsafe_allow_html=True,
        )
        if not no_master:
            k1, k2, k3, k4 = st.columns(4)
            kpi_box(k1, df_master["File Name"].nunique(), "Quotes", C_ORANGE)
            kpi_box(k2, df_exp_master["Service"].nunique(), "Services", C_DARK)
            kpi_box(k3, df_master["Vendor"].nunique(), "Vendors", C_MID)
            kpi_box(k4, df_master["Category"].nunique(), "Categories", C_GREY_DARK)
        else:
            st.warning("Master Catalog.xlsx not found.")

    with right:
        st.markdown(
            "<div class='bucket-header'>📊 DUMMY DATA (CSV — with prices)</div>",
            unsafe_allow_html=True,
        )
        if not no_dummy:
            k1, k2, k3, k4 = st.columns(4)
            kpi_box(k1, df_dummy["File Name"].nunique(), "Quotes", C_ORANGE)
            kpi_box(k2, df_exp_dummy["Service"].nunique(), "Services", C_DARK)
            kpi_box(k3, df_dummy["Vendor"].nunique(), "Vendors", C_MID)
            kpi_box(k4, df_dummy["Category"].nunique(), "Categories", C_GREY_DARK)
        else:
            st.warning("dummy_catalog.csv not found.")

    st.markdown("<br>", unsafe_allow_html=True)


def render_category_nav_strip(df_master, no_master):
    if no_master:
        return

    categories = sorted(
        [
            category
            for category in df_master["Category"].unique()
            if str(category).strip() not in ["", "nan"]
        ]
    )

    nav_cols = st.columns(len(categories) + 1)
    nav_cols[0].markdown(
        "<div style='font-size:0.72em;font-weight:700;color:#D04A02;letter-spacing:1px;"
        "text-transform:uppercase;padding-top:4px'>CATEGORIES</div>",
        unsafe_allow_html=True,
    )

    for idx, category in enumerate(categories):
        count = len(df_master[df_master["Category"] == category])
        nav_cols[idx + 1].markdown(
            "<div style='background:white;border:1px solid #E0E0E0;border-top:3px solid #D04A02;"
            "border-radius:4px;padding:8px 10px;text-align:center'>"
            "<div style='font-size:0.72em;font-weight:700;color:#2D2D2D;white-space:nowrap;"
            "overflow:hidden;text-overflow:ellipsis'>{}</div>"
            "<div style='font-size:1.1em;font-weight:800;color:#D04A02;font-family:Georgia,serif'>{}</div>"
            "<div style='font-size:0.63em;color:#7D7D7D;text-transform:uppercase;letter-spacing:0.5px'>"
            "quotes</div></div>".format(category, count),
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)


def render_upload_and_score_tab(df_master, df_exp_master, df_dummy, df_exp_dummy, no_dummy):
    st.markdown(
        "<div class='bucket-header'>📤 UPLOAD &amp; SCORE</div>",
        unsafe_allow_html=True,
    )

    benchmark_df = df_dummy if not no_dummy else df_master
    benchmark_exp_df = df_exp_dummy if not no_dummy else df_exp_master
    benchmark_label = "Dummy Data" if not no_dummy else "Master Catalog"

    if benchmark_df is None:
        st.info("No catalog loaded for benchmarking.")
        return

    insight("Benchmarking against: <b>{}</b>".format(benchmark_label))

    prefill_price = st.session_state.get("tab2_upload_price", 0.0)
    prefill_filename = st.session_state.get("tab2_upload_fname", "")

    if prefill_filename:
        st.markdown(
            "<div class='insight-box'>"
            "📎 From chat: <b>{}</b> — price: <b>{}</b>"
            "</div>".format(
                prefill_filename,
                fmt_currency(prefill_price) if prefill_price > 0 else "not found",
            ),
            unsafe_allow_html=True,
        )

    sec("STEP 1 — UPLOAD QUOTE FILE")
    uploaded_file = st.file_uploader(
        "Upload",
        type=["pdf", "xlsx", "xls", "docx"],
        label_visibility="collapsed",
        key="up_score_file",
    )

    new_price = 0.0
    uploaded_filename = ""

    if uploaded_file is not None:
        content = uploaded_file.read()
        ext = uploaded_file.name.rsplit(".", 1)[-1]
        uploaded_filename = uploaded_file.name

        st.success("Uploaded: **{}** ({} KB)".format(uploaded_filename, round(len(content) / 1024, 1)))

        sec("STEP 2 — EXTRACTED PRICE")
        with st.spinner("Extracting…"):
            result = extract_price_from_bytes(content, ext)
            new_price = result["price_num"]

        if new_price > 0:
            st.markdown(
                "<div class='scard scard-orange'>"
                "<div style='font-size:0.70em;font-weight:700;text-transform:uppercase;color:#D04A02'>"
                "Extracted Price</div>"
                "<div style='font-size:2.1em;font-weight:800;color:#D04A02;font-family:Georgia,serif'>"
                "{}</div></div>".format(fmt_currency(new_price)),
                unsafe_allow_html=True,
            )
        else:
            st.warning("Price not found automatically.")
            manual_price = st.number_input(
                "Enter price manually (USD)",
                min_value=0.0,
                step=100.0,
                value=0.0,
                key="manual_price_up",
            )
            if manual_price > 0:
                new_price = manual_price

    elif prefill_price > 0:
        new_price = prefill_price
        uploaded_filename = prefill_filename

    sec("STEP 3 — SELECT SERVICES & FILTERS")

    c1, c2, c3 = st.columns(3)

    with c1:
        selected_category = st.selectbox(
            "📂 Filter by Category",
            ["All"] + sorted(
                [
                    category
                    for category in benchmark_df["Category"].unique()
                    if str(category).strip() not in ["", "nan"]
                ]
            ),
            key="cat_up_score",
        )

    with c2:
        services = sorted(
            [
                service
                for service in benchmark_exp_df["Service"].unique()
                if str(service).strip() not in ["", "nan"]
            ]
        )
        service_search = st.text_input(
            "🔍 Filter services",
            placeholder="Search…",
            key="svc_srch_score",
        )
        if service_search:
            services = [service for service in services if service_search.lower() in service.lower()]

    with c3:
        selected_services = st.multiselect(
            "🛠 Select Services",
            options=services,
            key="new_svcs_score",
        )

    sec("STEP 4 — COMPARISON & VERDICT")

    if new_price <= 0 and not selected_services:
        st.info("Upload a file and select services to compare.")
        return

    candidate_df = (
        benchmark_exp_df[benchmark_exp_df["Service"].isin(selected_services)].copy()
        if selected_services
        else benchmark_exp_df.copy()
    )

    if selected_category != "All":
        candidate_df = candidate_df[candidate_df["Category"] == selected_category]

    if "Quoted Price" in candidate_df.columns:
        comparison_df = candidate_df.drop_duplicates(subset=["File Name", "Vendor"])[
            ["File Name", "Vendor", "Category", "Quoted Price"]
        ].copy()
    else:
        comparison_df = candidate_df.drop_duplicates(subset=["File Name", "Vendor"])[
            ["File Name", "Vendor", "Category"]
        ].copy()

    if comparison_df.empty:
        st.warning("No historical quotes found.")
        return

    historical_prices = []
    if "Quoted Price" in comparison_df.columns:
        for _, row in comparison_df.iterrows():
            quoted = parse_num(str(row.get("Quoted Price", "")).strip())
            if quoted > 0:
                historical_prices.append(quoted)

    if new_price > 0 and historical_prices:
        score, label, avg_hist, min_hist, max_hist = price_score(new_price, historical_prices)
        verdict_title, verdict_desc, verdict_color = get_verdict(score)

        css_key = "orange" if (score or 0) >= 70 else "mid" if (score or 0) >= 40 else "dark"
        backgrounds = {"orange": "#FFF5F0", "mid": "#F5F5F5", "dark": "#F0F0F0"}
        borders = {"orange": C_ORANGE, "mid": C_GREY_DARK, "dark": C_DARK}

        st.markdown(
            "<div style='background:{};border:2px solid {};border-radius:4px;padding:16px 20px;margin-bottom:16px'>"
            "<div style='font-size:1.2em;font-weight:700;color:{};font-family:Georgia,serif'>{}</div>"
            "<div style='font-size:0.87em;color:#4A4A4A;margin-top:5px'>{}</div>"
            "</div>".format(
                backgrounds[css_key],
                borders[css_key],
                borders[css_key],
                verdict_title,
                verdict_desc,
            ),
            unsafe_allow_html=True,
        )

        k1, k2, k3, k4 = st.columns(4)

        k1.markdown(
            "<div class='scard scard-orange'>"
            "<div style='font-size:0.67em;font-weight:700;text-transform:uppercase;color:#D04A02'>Score</div>"
            "<div style='font-size:2.1em;font-weight:800;color:#D04A02;font-family:Georgia,serif'>{}/100</div>"
            "<div style='font-size:0.73em;color:#7D7D7D;margin-top:3px'>vs {} historical</div>"
            "</div>".format(score if score is not None else "N/A", len(historical_prices)),
            unsafe_allow_html=True,
        )

        k2.markdown(
            "<div class='scard scard-dark'>"
            "<div style='font-size:0.67em;font-weight:700;text-transform:uppercase;color:#2D2D2D'>Your Price</div>"
            "<div style='font-size:2.1em;font-weight:800;color:#D04A02;font-family:Georgia,serif'>{}</div>"
            "</div>".format(fmt_currency(new_price)),
            unsafe_allow_html=True,
        )

        k3.markdown(
            "<div class='scard scard-dark'>"
            "<div style='font-size:0.67em;font-weight:700;text-transform:uppercase;color:#2D2D2D'>Market Average</div>"
            "<div style='font-size:2.1em;font-weight:800;color:#4A4A4A;font-family:Georgia,serif'>{}</div>"
            "<div style='font-size:0.72em;color:#7D7D7D;margin-top:3px'>min {} · max {}</div>"
            "</div>".format(
                fmt_currency(avg_hist),
                fmt_currency(min_hist),
                fmt_currency(max_hist),
            ),
            unsafe_allow_html=True,
        )

        k4.markdown(
            "<div class='scard scard-grey'>"
            "<div style='font-size:0.67em;font-weight:700;text-transform:uppercase;color:#7D7D7D'>vs Average</div>"
            "<div style='font-size:0.95em;font-weight:800;color:#4A4A4A;margin-top:8px'>{}</div>"
            "</div>".format(label),
            unsafe_allow_html=True,
        )

        st.markdown("<br>", unsafe_allow_html=True)
        sec("PRICE POSITIONING CHART")

        chart_rows = []
        if "Quoted Price" in comparison_df.columns:
            for _, row in comparison_df.iterrows():
                price_val = parse_num(str(row.get("Quoted Price", "")).strip())
                if price_val > 0:
                    chart_rows.append(
                        {
                            "Label": "{}/{}".format(row["Vendor"], str(row["File Name"])[:10]),
                            "Price": price_val,
                            "Type": "Historical",
                        }
                    )

        chart_rows.append({"Label": "★ YOUR QUOTE", "Price": new_price, "Type": "New"})
        chart_df = pd.DataFrame(chart_rows).sort_values("Price")

        bar_colors = [C_ORANGE if t == "New" else C_DARK for t in chart_df["Type"]]

        fig = go.Figure(
            go.Bar(
                x=chart_df["Label"],
                y=chart_df["Price"],
                marker_color=bar_colors,
                marker_line_width=0,
                text=chart_df["Price"].apply(fmt_currency),
                textposition="outside",
            )
        )

        fig.add_hline(
            y=avg_hist,
            line_dash="dash",
            line_color=C_GREY_DARK,
            line_width=2,
            annotation_text="Avg: {}".format(fmt_currency(avg_hist)),
            annotation_position="top right",
        )

        fig.update_layout(
            height=380,
            plot_bgcolor=CBG,
            paper_bgcolor=CBG,
            margin=dict(l=5, r=10, t=20, b=10),
            font=CFONT,
            yaxis=dict(title="Price (USD)", showgrid=True, gridcolor=C_GREY_LITE, zeroline=False),
            xaxis=dict(tickangle=-25),
            bargap=0.3,
            showlegend=False,
        )
        st.plotly_chart(fig, use_container_width=True)

        pct_vs_avg = round((new_price - avg_hist) / avg_hist * 100, 1) if avg_hist > 0 else 0
        insight(
            "Your quote <b>{}</b> is <b>{}% {}</b> market avg <b>{}</b>. "
            "Range: <b>{}</b>–<b>{}</b>.".format(
                fmt_currency(new_price),
                abs(pct_vs_avg),
                "below" if pct_vs_avg < 0 else "above",
                fmt_currency(avg_hist),
                fmt_currency(min_hist),
                fmt_currency(max_hist),
            )
        )
    else:
        st.info("No historical price data. Select more services or use Dummy Data tab.")


def render_data_table_tab(df_master, df_dummy):
    st.markdown(
        "<div class='bucket-header'>📄 DATA TABLE</div>",
        unsafe_allow_html=True,
    )

    source_choice = st.radio(
        "Data source",
        ["Master Catalog", "Dummy Data"],
        horizontal=True,
        key="dt_src",
    )

    selected_df = df_master if source_choice == "Master Catalog" else df_dummy
    selected_df = selected_df if selected_df is not None else pd.DataFrame()

    c1, c2, c3 = st.columns(3)

    with c1:
        category_options = (
            ["All"] + sorted(
                [
                    category
                    for category in selected_df.get("Category", pd.Series(dtype=str)).unique()
                    if str(category).strip() not in ["", "nan"]
                ]
            )
            if not selected_df.empty
            else ["All"]
        )
        selected_category = st.selectbox("📂 Category", category_options, key="dt_cat")

    with c2:
        vendor_pool = selected_df if selected_category == "All" else selected_df[selected_df["Category"] == selected_category]
        vendor_options = (
            ["All"] + sorted(
                [
                    vendor
                    for vendor in vendor_pool.get("Vendor", pd.Series(dtype=str)).unique()
                    if str(vendor).strip() not in ["", "nan"]
                ]
            )
            if not vendor_pool.empty
            else ["All"]
        )
        selected_vendor = st.selectbox("🏢 Vendor", vendor_options, key="dt_ven")

    with c3:
        search_text = st.text_input("🔍 Search", placeholder="File name or comments…", key="dt_srch")

    filtered_df = selected_df.copy()
    if not filtered_df.empty:
        if selected_category != "All":
            filtered_df = filtered_df[filtered_df["Category"] == selected_category]
        if selected_vendor != "All":
            filtered_df = filtered_df[filtered_df["Vendor"] == selected_vendor]
        if search_text:
            mask = (
                filtered_df["File Name"].str.contains(search_text, case=False, na=False)
                | filtered_df["Comments"].str.contains(search_text, case=False, na=False)
            )
            filtered_df = filtered_df[mask]

        st.markdown(
            "<div style='font-size:0.82em;color:#7D7D7D;margin:8px 0'>"
            "Showing <b>{}</b> of <b>{}</b> records</div>".format(
                len(filtered_df),
                len(selected_df),
            ),
            unsafe_allow_html=True,
        )

        st.dataframe(
            filtered_df.drop(columns=["Services List", "Hyperlink"], errors="ignore"),
            use_container_width=True,
            height=520,
        )
    else:
        st.info("No data available.")


def render_upload_catalog_tab():
    st.markdown(
        "<div class='bucket-header'>🗂 UPLOAD CATALOG</div>",
        unsafe_allow_html=True,
    )

    insight("Upload a new catalog to replace the current one. Supports Excel (.xlsx) and CSV formats.")

    catalog_file = st.file_uploader(
        "Upload",
        type=["xlsx", "xls", "csv"],
        label_visibility="collapsed",
        key="catalog_upload",
    )

    if catalog_file is None:
        return

    file_bytes = catalog_file.read()
    filename = catalog_file.name

    with st.spinner("Analysing…"):
        from io import BytesIO

        ext = filename.rsplit(".", 1)[-1].lower()

        try:
            if ext in ("xlsx", "xls"):
                df_new = pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
            else:
                df_new = pd.read_csv(BytesIO(file_bytes))

            df_new.columns = [str(col).strip() for col in df_new.columns]
            df_new = normalize_columns(df_new)
            df_new = clean_df(df_new)
            df_new["Hyperlink"] = ""
            df_new, df_exp_new = explode_services(df_new)

            st.success(
                "✅ **{}** rows · **{}** vendors · **{}** categories".format(
                    len(df_new),
                    df_new["Vendor"].nunique(),
                    df_new["Category"].nunique(),
                )
            )

            st.dataframe(
                df_new.drop(columns=["Services List", "Hyperlink"], errors="ignore").head(20),
                use_container_width=True,
                height=280,
            )

            if st.button("✅ Apply as Master Catalog", type="primary"):
                st.session_state["uploaded_catalog_df"] = df_new
                st.session_state["uploaded_catalog_exp"] = df_exp_new
                st.success("✅ Applied!")
                st.rerun()

        except Exception as exc:
            st.error("❌ {}".format(exc))


def render_vendor_analysis_tab(df_dummy, df_exp_dummy, vcmap_dummy, no_dummy):
    st.markdown(
        "<div class='bucket-header'>🔍 VENDOR ANALYSIS — DUMMY DATA (Full Price Analysis)</div>",
        unsafe_allow_html=True,
    )

    if no_dummy:
        st.info("Dummy data not available.")
        return

    sec("VENDOR PRICE SUMMARY", "Based on dummy data with actual prices")

    working_df = df_dummy.copy()
    if "Quoted Price" in working_df.columns:
        working_df["Quoted Price"] = pd.to_numeric(working_df["Quoted Price"], errors="coerce")
        vendor_table = (
            working_df.groupby("Vendor")["Quoted Price"]
            .agg(["mean", "min", "max", "count"])
            .reset_index()
        )
    else:
        vendor_table = pd.DataFrame()

    if vendor_table.empty:
        st.info("No vendor pricing data available.")
        return

    vendor_table.columns = ["Vendor", "Average", "Min", "Max", "Quotes"]
    vendor_table = vendor_table.sort_values("Average")
    overall_avg = vendor_table["Average"].mean()

    k1, k2, k3, k4 = st.columns(4)
    kpi_box(k1, len(df_dummy), "Total Quotes", C_ORANGE)
    kpi_box(k2, df_dummy["Vendor"].nunique(), "Vendors", C_DARK)
    kpi_box(k3, fmt_currency(overall_avg), "Overall Avg Quote", C_MID)
    kpi_box(k4, df_exp_dummy["Service"].nunique(), "Services", C_GREY_DARK)

    st.markdown("<br>", unsafe_allow_html=True)

    bar_colors = [
        C_ORANGE if idx == 0 else C_DARK if idx == len(vendor_table) - 1 else C_GREY_DARK
        for idx in range(len(vendor_table))
    ]

    fig = go.Figure(
        go.Bar(
            x=vendor_table["Vendor"],
            y=vendor_table["Average"],
            marker_color=bar_colors,
            marker_line_width=0,
            text=vendor_table["Average"].apply(fmt_currency),
            textposition="outside",
        )
    )
    fig.add_hline(
        y=overall_avg,
        line_dash="dash",
        line_color=C_MID,
        line_width=2,
        annotation_text="Avg: {}".format(fmt_currency(overall_avg)),
        annotation_position="top right",
    )
    pwc_bar(fig, "Average Quote per Vendor", height=360)
    fig.update_xaxes(tickangle=-20)
    st.plotly_chart(fig, use_container_width=True)

    sec("VENDOR RANKING TABLE")

    table_rows = [
        "<table class='comp-table'><thead><tr>"
        "<th>Rank</th><th>Vendor</th><th>Quotes</th><th>Avg</th><th>Min</th><th>Max</th><th>vs Avg</th><th>Verdict</th>"
        "</tr></thead><tbody>"
    ]

    for rank, (_, row) in enumerate(vendor_table.iterrows(), start=1):
        bg = "white" if rank % 2 == 0 else "#F8F8F8"
        vendor_color = vcmap_dummy.get(row["Vendor"], C_DARK)
        pct = round((row["Average"] - overall_avg) / overall_avg * 100, 1) if overall_avg > 0 else 0
        pct_color = C_ORANGE if pct < -5 else C_DARK if pct > 5 else C_GREY_DARK
        pct_text = "{}% below".format(abs(pct)) if pct < 0 else "{}% above".format(abs(pct)) if pct > 0 else "At avg"
        overall_verdict = (
            ("✅ COMPETITIVE", C_ORANGE)
            if pct < -10
            else ("🔴 EXPENSIVE", C_DARK)
            if pct > 10
            else ("🟡 AVERAGE", C_GREY_DARK)
        )
        medal = "🥇" if rank == 1 else "🥈" if rank == 2 else "🥉" if rank == 3 else str(rank)

        table_rows.append(
            "<tr style='background:{}'>"
            "<td style='text-align:center;font-size:1.05em'>{}</td>"
            "<td>{}</td>"
            "<td style='text-align:center;font-weight:700'>{}</td>"
            "<td style='font-family:monospace;font-weight:700;color:#D04A02'>{}</td>"
            "<td style='font-family:monospace;color:#D04A02'>{}</td>"
            "<td style='font-family:monospace;color:#2D2D2D'>{}</td>"
            "<td style='color:{}'>{}</td>"
            "<td style='color:{};font-weight:700'>{}</td>"
            "</tr>".format(
                bg,
                medal,
                vpill(row["Vendor"], vendor_color),
                int(row["Quotes"]),
                fmt_currency(row["Average"]),
                fmt_currency(row["Min"]),
                fmt_currency(row["Max"]),
                pct_color,
                pct_text,
                overall_verdict[1],
                overall_verdict[0],
            )
        )

    table_rows.append("</tbody></table>")
    render_html_table(table_rows)

    st.markdown("<br>", unsafe_allow_html=True)
    sec("PER-SERVICE PRICE BENCHMARKING", "Services with multiple vendor quotes")

    service_rows = []
    for _, row in df_dummy.iterrows():
        raw_services = (
            str(row.get("Comments", ""))
            .replace("\\n", "\n")
            .replace("\r\n", "\n")
            .replace("\r", "\n")
        )
        services = [
            service.strip()
            for service in raw_services.split("\n")
            if service.strip() and service.strip() not in ["nan", "None", ""]
        ]
        if not services:
            services = [raw_services.strip()]

        for service in services:
            service_rows.append(
                {
                    "Service": service,
                    "Vendor": row["Vendor"],
                    "Price": parse_num(str(row.get("Quoted Price", "")).strip()),
                }
            )

    service_df = pd.DataFrame(service_rows)
    service_df = service_df[service_df["Price"] > 0]

    service_vendor_counts = service_df.groupby("Service")["Vendor"].nunique()
    multi_vendor_services = service_vendor_counts[service_vendor_counts > 1].index.tolist()

    if not multi_vendor_services:
        st.info("No multi-vendor services.")
    else:
        insight("<b>{}</b> services with multi-vendor quotes.".format(len(multi_vendor_services)))

        for service in sorted(multi_vendor_services)[:10]:
            service_slice = service_df[service_df["Service"] == service].sort_values("Price")
            min_price = service_slice["Price"].min()
            max_price = service_slice["Price"].max()
            avg_price = service_slice["Price"].mean()
            best_vendor = service_slice.loc[service_slice["Price"].idxmin(), "Vendor"]
            worst_vendor = service_slice.loc[service_slice["Price"].idxmax(), "Vendor"]
            spread = round((max_price - min_price) / min_price * 100, 1) if min_price > 0 else 0

            st.markdown(
                "<div style='background:white;border-left:4px solid {};padding:10px 14px;border-radius:2px;margin:10px 0;font-weight:700;font-size:0.88em'>"
                "{} · {} vendors · spread {}% · best: {} @ {}"
                "</div>".format(
                    C_ORANGE,
                    service,
                    service_slice["Vendor"].nunique(),
                    spread,
                    best_vendor,
                    fmt_currency(min_price),
                ),
                unsafe_allow_html=True,
            )

            bar_colors = [
                C_ORANGE if vendor == best_vendor else C_DARK if vendor == worst_vendor else C_GREY_DARK
                for vendor in service_slice["Vendor"]
            ]

            fig = go.Figure(
                go.Bar(
                    x=service_slice["Vendor"],
                    y=service_slice["Price"],
                    marker_color=bar_colors,
                    marker_line_width=0,
                    text=service_slice["Price"].apply(fmt_currency),
                    textposition="outside",
                )
            )

            fig.add_hline(
                y=avg_price,
                line_dash="dash",
                line_color=C_MID,
                line_width=1.5,
                annotation_text="Avg: {}".format(fmt_currency(avg_price)),
                annotation_position="top right",
            )

            fig.update_layout(
                height=240,
                plot_bgcolor=CBG,
                paper_bgcolor=CBG,
                margin=dict(l=5, r=10, t=12, b=8),
                font=CFONT,
                yaxis=dict(showgrid=True, gridcolor=C_GREY_LITE, zeroline=False),
                bargap=0.4,
                showlegend=False,
            )
            st.plotly_chart(fig, use_container_width=True)

    st.markdown("<br>", unsafe_allow_html=True)
    st.download_button(
        "📥 Download Dummy Data CSV",
        data=df_dummy.to_csv(index=False),
        file_name="dummy_analysis.csv",
        mime="text/csv",
        type="primary",
    )


# ============================================================
# APP BOOTSTRAP
# ============================================================

init_session_defaults()
ensure_dummy_data()

df_master, df_exp_master = load_master_catalog()
df_dummy, df_exp_dummy = load_dummy_data()

df_master, df_exp_master = apply_uploaded_catalog_if_present(df_master, df_exp_master)

NO_MASTER = df_master is None or df_master.empty
NO_DUMMY = df_dummy is None or df_dummy.empty

vcmap_master, vcmap_dummy = vendor_color_maps(df_master, df_dummy)
vendor_color_map = vcmap_master if not NO_MASTER else vcmap_dummy

render_main_header()
render_top_kpis(df_master, df_exp_master, df_dummy, df_exp_dummy, NO_MASTER, NO_DUMMY)
render_category_nav_strip(df_master, NO_MASTER)


# ============================================================
# MAIN TABS
# ============================================================

(
    tab_mc_ov,
    tab_mc_bv,
    tab_dd_ov,
    tab_dd_bv,
    tab_upload,
    tab_data,
    tab_upload_cat,
    tab_vendor,
) = st.tabs(
    [
        "📁 MC — Catalog Overview",
        "📁 MC — Browse & Verdict",
        "📊 DD — Catalog Overview",
        "📊 DD — Browse & Verdict",
        "📤 Upload & Score",
        "📄 Data Table",
        "🗂 Upload Catalog",
        "🔍 Vendor Analysis",
    ]
)


with tab_mc_ov:
    st.markdown(
        "<div class='bucket-header'>📁 MASTER CATALOG — CATALOG OVERVIEW</div>",
        unsafe_allow_html=True,
    )
    if NO_MASTER:
        st.info("Master Catalog.xlsx not found.")
    else:
        render_catalog_overview(df_master, df_exp_master, label="Master Catalog")


with tab_mc_bv:
    st.markdown(
        "<div class='bucket-header'>"
        "📁 MASTER CATALOG — BROWSE &amp; VERDICT"
        "<span style='font-size:0.8em;opacity:0.7;margin-left:12px'>"
        "No prices available (SharePoint files)</span>"
        "</div>",
        unsafe_allow_html=True,
    )
    if NO_MASTER:
        st.info("Master Catalog.xlsx not found.")
    else:
        render_browse_verdict(
            df_master=df_master,
            df_exploded=df_exp_master,
            vcmap=vcmap_master,
            label="Master Catalog",
            has_prices=False,
            chat_key_suffix="mc",
        )


with tab_dd_ov:
    st.markdown(
        "<div class='bucket-header'>"
        "📊 DUMMY DATA — CATALOG OVERVIEW"
        "<span style='font-size:0.8em;opacity:0.7;margin-left:12px'>"
        "Full price analysis available</span>"
        "</div>",
        unsafe_allow_html=True,
    )
    if NO_DUMMY:
        st.info("dummy_catalog.csv not found.")
    else:
        render_catalog_overview(df_dummy, df_exp_dummy, label="Dummy Data")


with tab_dd_bv:
    st.markdown(
        "<div class='bucket-header'>"
        "📊 DUMMY DATA — BROWSE &amp; VERDICT"
        "<span style='font-size:0.8em;opacity:0.7;margin-left:12px'>"
        "✅ Full price analysis · Scores · Verdicts"
        "</span></div>",
        unsafe_allow_html=True,
    )
    if NO_DUMMY:
        st.info("dummy_catalog.csv not found.")
    else:
        render_browse_verdict(
            df_master=df_dummy,
            df_exploded=df_exp_dummy,
            vcmap=vcmap_dummy,
            label="Dummy Data",
            has_prices=True,
            chat_key_suffix="dd",
        )


with tab_upload:
    render_upload_and_score_tab(df_master, df_exp_master, df_dummy, df_exp_dummy, NO_DUMMY)


with tab_data:
    render_data_table_tab(df_master, df_dummy)


with tab_upload_cat:
    render_upload_catalog_tab()


with tab_vendor:
    render_vendor_analysis_tab(df_dummy, df_exp_dummy, vcmap_dummy, NO_DUMMY)
